import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, PanedWindow
import json
import os
import threading
import time
from datetime import datetime, timedelta, date
import webbrowser
import pandas as pd
import sqlite3
import xlwings as xw
import csv
from datetime import datetime as dt

try:
    from kiteconnect import KiteConnect
except ImportError:
    print("Please install kiteconnect: pip install kiteconnect")
    exit()

import openpyxl
import os

print_debug = False
FILE_NAME = 'MCX_Trading_Platform_Data.xlsx'

def create_initial_file():
    """
    Function to create a new Excel file and add some initial data.
    """
    if print_debug:
        print(f"--- Creating initial file: {FILE_NAME} ---")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Future Readings"
    
    # Add headers
    sheet['A1'] = 'Date'
    sheet['B1'] = 'Time'
    sheet['C1'] = 'Value'
    
    # Save the workbook
    workbook.save(FILE_NAME)
    if print_debug:
        print(f"Created and saved {FILE_NAME}\n")

def update_existing_file(value_price):
    """
    Function to open an existing Excel file (created by another function), 
    modify a cell, and save the changes.
    """
    if print_debug:
        print(f"--- Updating file: {FILE_NAME} ---")
    if not os.path.exists(FILE_NAME):
        print(f"Error: {FILE_NAME} not found. Run create_initial_file() first.")
        return

    # Load the existing workbook
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook['Future Readings'] # Access the specific sheet by name
    
    # Add a new row of data (optional)
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=dt.now().date())
    sheet.cell(row=next_row, column=2, value=dt.now().time())
    sheet.cell(row=next_row, column=3, value=value_price)
    
    # Save the workbook (overwrites the old one)
    workbook.save(FILE_NAME)
    if print_debug:
        print(f"Updated and saved {FILE_NAME}\n")


class ZerodhaTradingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("MCX Trading Platform - Entry/Exit Signals")
        self.root.geometry("1400x900")
        
        # Make main window resizable and draggable
        self.root.resizable(True, True)
        
        # Initialize variables
        self.kite = None
        self.is_logged_in = False
        self.api_key = ""
        self.access_token = ""
        self.live_data = {}
        self.positions = {}
        self.orders = {}
        self.profit_target = 0
        self.total_pnl = 0
        self.instruments_df = None
        
        # Live data flags
        self.live_data_running = False
        self.futures_data_running = False
        self.options_data_running = False
        
        # Month comparison
        self.month_comparison_running = False
        self.current_month_contract = None
        self.next_month_contract = None
        self.comparison_popup = None
        
        # PREVIOUS DAY CLOSING PRICES storage
        self.previous_day_close_prices = {}
        self.month_comparison_prices = {}
        
        # Daily performance tracking
        self.daily_performance_db = "daily_performance.db"
        
        # NEW: Triggered popup variables
        self.triggered_popup = None
        self.last_trigger_time = None
        self.trigger_cooldown = 600  # seconds between triggers
        self.trigger_threshold = 0.2  # percentage threshold difference
        
        # NEW: Price difference popup
        self.price_diff_popup = None
        
        # NEW: Entry/Exit popup variables
        self.entry_exit_popup = None
        self.last_entry_exit_trigger_time = None
        self.entry_exit_cooldown = 600  # 1 minutes cooldown
        self.entry_threshold = -10.0  # Less than -10 for entry
        self.exit_threshold = 10.0    # More than +6 for exit
        
        # NEW: Trading variables
        self.current_quantity = 1  # Default quantity
        self.order_history = []
        self.pending_orders = []
        
        # NEW: Calendar spread trading variables
        self.spread_orders = []
        self.spread_position = None
        self.spread_quantity = 1
        self.auto_spread_trading = False
        
        # NEW: Auto-exit variables
        self.auto_exit_enabled = False
        self.auto_exit_profit_target = 0.0  # Default profit target in rupees
        self.auto_exit_stop_loss = 0.0  # Default stop loss in rupees
        self.trade_start_price = 0.0
        self.trade_direction = None  # "BUY" or "SELL"
        self.current_position = None
        self.auto_exit_running = False
        
        # Load credentials
        self.load_credentials()
        
        # Initialize database for daily tracking
        self.init_daily_performance_db()
        
        # Setup GUI with resizable frames
        self.setup_gui_with_resizable_frames()
        
        # Auto login if credentials exist
        if hasattr(self, 'api_key') and hasattr(self, 'access_token') and self.api_key and self.access_token:
            self.root.after(1000, self.auto_login)

    def setup_gui_with_resizable_frames(self):
        """Setup the main GUI with resizable frames"""
        # Create main container with grid
        main_container = ttk.Frame(self.root)
        main_container.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Configure grid weights for resizing
        main_container.grid_columnconfigure(0, weight=3)  # Notebook area
        main_container.grid_columnconfigure(1, weight=1)  # Log area
        main_container.grid_rowconfigure(0, weight=1)
        
        # Create notebook for tabs on left
        notebook_frame = ttk.Frame(main_container)
        notebook_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 5))
        
        notebook_frame.grid_columnconfigure(0, weight=1)
        notebook_frame.grid_rowconfigure(0, weight=1)
        
        notebook = ttk.Notebook(notebook_frame)
        notebook.grid(row=0, column=0, sticky='nsew')
        
        # Log message area on right
        log_frame = ttk.LabelFrame(main_container, text="Log Messages")
        log_frame.grid(row=0, column=1, sticky='nsew')
        
        log_frame.grid_columnconfigure(0, weight=1)
        log_frame.grid_rowconfigure(0, weight=1)
        log_frame.grid_rowconfigure(1, weight=0)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=40)
        self.log_text.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        
        # Add clear button at bottom
        button_frame = ttk.Frame(log_frame)
        button_frame.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        
        ttk.Button(button_frame, text="Clear Log", 
                  command=self.clear_log).pack(side='right', padx=5)
        
        # Login Tab
        self.setup_login_tab(notebook)
        
        # Month Comparison Tab (Updated for Previous Day Close)
        self.setup_month_comparison_tab(notebook)
        
        # Trading Tab
        self.setup_trading_tab(notebook)
        
        # NEW: Calendar Spread Trading Tab
        self.setup_calendar_spread_tab(notebook)
        
        # NEW: Auto Exit Tab
        self.setup_auto_exit_tab(notebook)

    def clear_log(self):
        """Clear the log messages"""
        self.log_text.delete(1.0, tk.END)

    def load_credentials(self):
        """Load API credentials from file"""
        try:
            if os.path.exists('zerodha_credentials.json'):
                with open('zerodha_credentials.json', 'r') as f:
                    creds = json.load(f)
                    self.api_key = creds.get('api_key', '')
                    self.access_token = creds.get('access_token', '')
                    self.log_message("Credentials loaded successfully")
        except Exception as e:
            self.log_message(f"Error loading credentials: {e}")

    def save_credentials(self):
        """Save API credentials to file"""
        try:
            creds = {
                'api_key': self.api_key,
                'access_token': self.access_token
            }
            with open('zerodha_credentials.json', 'w') as f:
                json.dump(creds, f, indent=4)
            self.log_message("Credentials saved successfully")
        except Exception as e:
            self.log_message(f"Error saving credentials: {e}")

    def setup_login_tab(self, notebook):
        """Setup login tab"""
        login_frame = ttk.Frame(notebook)
        notebook.add(login_frame, text="Login")
        
        # Use grid for resizable layout
        login_frame.grid_columnconfigure(0, weight=1)
        login_frame.grid_rowconfigure(0, weight=1)
        login_frame.grid_rowconfigure(1, weight=1)
        
        # Create main container with scrollbar if needed
        container = ttk.Frame(login_frame)
        container.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
        
        # Top section for input fields
        input_frame = ttk.LabelFrame(container, text="Login Credentials")
        input_frame.pack(fill='both', expand=True,  pady=(0, 10))
        
        # API Key
        ttk.Label(input_frame, text="API Key:").grid(row=0, column=0, padx=10, pady=10, sticky='w')
        self.api_key_entry = ttk.Entry(input_frame, width=40)
        self.api_key_entry.grid(row=0, column=1, padx=10, pady=10)
        if hasattr(self, 'api_key'):
            self.api_key_entry.insert(0, self.api_key)
        
        # API Secret
        ttk.Label(input_frame, text="API Secret:").grid(row=1, column=0, padx=10, pady=10, sticky='w')
        self.api_secret_entry = ttk.Entry(input_frame, width=40, show='*')
        self.api_secret_entry.grid(row=1, column=1, padx=10, pady=10)
        
        # Request Token
        ttk.Label(input_frame, text="Request Token:").grid(row=2, column=0, padx=10, pady=10, sticky='w')
        self.request_token_entry = ttk.Entry(input_frame, width=40)
        self.request_token_entry.grid(row=2, column=1, padx=10, pady=10)
        
        # Buttons frame
        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=3, column=0, columnspan=2, pady=15)
        
        ttk.Button(button_frame, text="Generate Login URL", 
                  command=self.generate_login_url).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Login", 
                  command=self.manual_login).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Auto Login", 
                  command=self.auto_login).pack(side='left', padx=5)
        
        # Status
        self.login_status = ttk.Label(input_frame, text="Not Logged In", foreground='red')
        self.login_status.grid(row=4, column=0, columnspan=2, padx=10, pady=10)
        
        # Bottom section for instructions
        instructions_frame = ttk.LabelFrame(container, text="Instructions")
        instructions_frame.pack(fill='both', expand=True)
        
        instructions = """
        Instructions:
        1. Enter your API Key and Secret (get from Zerodha developer console)
        2. Click 'Generate Login URL' and login to Zerodha
        3. After login, copy the request token from URL and paste above
        4. Click 'Login' to authenticate
        5. Use 'Auto Login' for future sessions
        
        Note: Your credentials are saved locally for auto-login.
        """
        ttk.Label(instructions_frame, text=instructions, justify='left').pack(padx=10, pady=10, fill='both', expand=True)

    def setup_month_comparison_tab(self, notebook):
        """Setup month comparison tab using PREVIOUS DAY CLOSING prices with resizable panes"""
        month_frame = ttk.Frame(notebook)
        notebook.add(month_frame, text="📅 Month Comparison (Prev Day Close)")
        
        # Configure grid for resizable layout
        month_frame.grid_columnconfigure(0, weight=1)  # Configuration column
        month_frame.grid_columnconfigure(1, weight=3)  # Display column
        month_frame.grid_rowconfigure(0, weight=1)
        
        # Left frame - Configuration and controls
        left_frame = ttk.Frame(month_frame)
        left_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 5))
        
        # Right frame - Comparison display and signals
        right_frame = ttk.Frame(month_frame)
        right_frame.grid(row=0, column=1, sticky='nsew')
        
        # Configure left frame grid
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_rowconfigure(0, weight=1)  # Configuration
        left_frame.grid_rowconfigure(1, weight=1)  # Trigger settings
        left_frame.grid_rowconfigure(2, weight=1)  # Previous day close
        left_frame.grid_rowconfigure(3, weight=1)  # Entry/exit settings
        left_frame.grid_rowconfigure(4, weight=1)  # Controls
        left_frame.grid_rowconfigure(5, weight=1)  # History (takes remaining space)
        
        # Configuration Frame
        config_frame = ttk.LabelFrame(left_frame, text="Month Comparison Configuration")
        config_frame.grid(row=0, column=0, sticky='nsew', pady=(0, 5))
        
        config_frame.grid_columnconfigure(1, weight=1)
        
        # Commodity selection
        ttk.Label(config_frame, text="Commodity:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.month_commodity = ttk.Combobox(config_frame, values=["GOLD", "SILVER", "CRUDEOIL", "NATURALGAS", "COPPER", "LEAD", "ZINC"])
        self.month_commodity.grid(row=0, column=1, padx=5, pady=5, sticky='nsew')
        self.month_commodity.set("GOLD")
        
        # Load contracts button
        ttk.Button(config_frame, text="Load Current & Next Month", 
                  command=self.load_month_contracts).grid(row=1, column=0, columnspan=2, pady=10)
        
        # Trigger settings frame
        trigger_frame = ttk.LabelFrame(left_frame, text="Trigger Settings")
        trigger_frame.grid(row=1, column=0, sticky='nsew', pady=5)
        
        trigger_frame.grid_columnconfigure(1, weight=1)
        
        # Trigger threshold
        ttk.Label(trigger_frame, text="Trigger Threshold (%):").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.trigger_threshold_var = tk.StringVar(value="0.2")
        self.trigger_threshold_entry = ttk.Entry(trigger_frame, textvariable=self.trigger_threshold_var, width=10)
        self.trigger_threshold_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        ttk.Label(trigger_frame, text="% difference").grid(row=0, column=2, padx=5, pady=5)
        
        # Cooldown period
        ttk.Label(trigger_frame, text="Cooldown (sec):").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        self.cooldown_var = tk.StringVar(value="600")
        self.cooldown_entry = ttk.Entry(trigger_frame, textvariable=self.cooldown_var, width=10)
        self.cooldown_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        # Test trigger button
        ttk.Button(trigger_frame, text="Test Trigger Popup", 
                  command=self.test_triggered_popup).grid(row=2, column=0, columnspan=3, pady=10)
        
        # PREVIOUS DAY CLOSE settings
        time_frame = ttk.LabelFrame(left_frame, text="Previous Day Close Settings")
        time_frame.grid(row=2, column=0, sticky='nsew', pady=5)
        
        ttk.Label(time_frame, text="Using Previous Trading Day Close:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        
        # Fetch previous day close button
        ttk.Button(time_frame, text="Fetch Previous Day Close", 
                  command=self.fetch_previous_day_closes).grid(row=1, column=0, columnspan=2, pady=5)
        
        ttk.Label(time_frame, text="Manual Previous Close:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        ttk.Button(time_frame, text="Set Manually", 
                  command=self.set_manual_previous_close).grid(row=2, column=1, padx=5, pady=5)
        
        # Entry/Exit Settings Frame
        entry_exit_frame = ttk.LabelFrame(left_frame, text="Entry/Exit Settings")
        entry_exit_frame.grid(row=3, column=0, sticky='nsew', pady=5)
        
        entry_exit_frame.grid_columnconfigure(1, weight=1)
        
        # Entry threshold (less than -10)
        ttk.Label(entry_exit_frame, text="Entry Threshold (₹):").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.entry_threshold_var = tk.StringVar(value="-10.0")
        self.entry_threshold_entry = ttk.Entry(entry_exit_frame, textvariable=self.entry_threshold_var, width=10)
        self.entry_threshold_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        ttk.Label(entry_exit_frame, text="Less than").grid(row=0, column=2, padx=5, pady=5)
        
        # Exit threshold (more than +6)
        ttk.Label(entry_exit_frame, text="Exit Threshold (₹):").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        self.exit_threshold_var = tk.StringVar(value="10.0")
        self.exit_threshold_entry = ttk.Entry(entry_exit_frame, textvariable=self.exit_threshold_var, width=10)
        self.exit_threshold_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        ttk.Label(entry_exit_frame, text="More than").grid(row=1, column=2, padx=5, pady=5)
        
        # Entry/Exit cooldown
        ttk.Label(entry_exit_frame, text="Cooldown (min):").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        self.entry_exit_cooldown_var = tk.StringVar(value="5")
        self.entry_exit_cooldown_entry = ttk.Entry(entry_exit_frame, textvariable=self.entry_exit_cooldown_var, width=10)
        self.entry_exit_cooldown_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        ttk.Label(entry_exit_frame, text="minutes").grid(row=2, column=2, padx=5, pady=5)
        
        # Test Entry/Exit button
        ttk.Button(entry_exit_frame, text="Test Entry/Exit Popup", 
                  command=self.test_entry_exit_popup).grid(row=3, column=0, columnspan=3, pady=10)
        
        # Control buttons frame
        control_frame = ttk.LabelFrame(left_frame, text="Controls")
        control_frame.grid(row=4, column=0, sticky='nsew', pady=5)
        
        control_buttons = ttk.Frame(control_frame)
        control_buttons.pack(fill='both', expand=True,  padx=5, pady=5)
        
        self.start_month_btn = ttk.Button(control_buttons, text="Start Month Comparison", 
                                         command=self.start_month_comparison)
        self.start_month_btn.pack(side='left', padx=2, pady=5, fill='x', expand=True)
        
        self.stop_month_btn = ttk.Button(control_buttons, text="Stop Comparison", 
                                        command=self.stop_month_comparison, state='disabled')
        self.stop_month_btn.pack(side='left', padx=2, pady=5, fill='x', expand=True)
        
        # Additional buttons frame
        more_buttons = ttk.Frame(control_frame)
        more_buttons.pack(fill='both', expand=True,  padx=5, pady=5)
        
        ttk.Button(more_buttons, text="Show Comparison Popup", 
                  command=self.show_comparison_popup).pack(side='left', padx=2, pady=2, fill='x', expand=True)
        
        ttk.Button(more_buttons, text="Show Price Diff Popup", 
                  command=self.show_price_difference_popup).pack(side='left', padx=2, pady=2, fill='x', expand=True)
        
        # Historical Performance Frame (takes remaining space)
        history_frame = ttk.LabelFrame(left_frame, text="Historical Performance (Last 7 Days)")
        history_frame.grid(row=5, column=0, sticky='nsew', pady=(5, 0))
        
        history_frame.grid_columnconfigure(0, weight=1)
        history_frame.grid_rowconfigure(0, weight=1)
        
        self.history_text = scrolledtext.ScrolledText(history_frame, height=10)
        self.history_text.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        self.history_text.insert(tk.END, "Load contracts and start monitoring to see history")
        
        # RIGHT FRAME - Display sections
        # Configure right frame grid
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(0, weight=3)  # Comparison display
        right_frame.grid_rowconfigure(1, weight=1)  # Total changes
        right_frame.grid_rowconfigure(2, weight=1)  # Price difference
        right_frame.grid_rowconfigure(3, weight=1)  # Signal display
        right_frame.grid_rowconfigure(4, weight=1)  # Status
        
        # Top section: Current vs Next Month Comparison
        display_frame = ttk.LabelFrame(right_frame, text="Current vs Next Month Comparison (vs Prev Day Close)")
        display_frame.grid(row=0, column=0, sticky='nsew', pady=(0, 5))
        
        display_frame.grid_columnconfigure(0, weight=1)  # Current month
        display_frame.grid_columnconfigure(1, weight=1)  # VS separator
        display_frame.grid_columnconfigure(2, weight=1)  # Next month
        display_frame.grid_columnconfigure(3, weight=1)  # Smiley
        display_frame.grid_rowconfigure(0, weight=1)
        
        # Current month frame
        current_frame = ttk.LabelFrame(display_frame, text="Current Month")
        current_frame.grid(row=0, column=0, sticky='nsew', padx=(5, 2), pady=5)
        
        current_frame.grid_columnconfigure(0, weight=1)
        
        self.current_contract_label = ttk.Label(current_frame, text="--", font=('Arial', 10))
        self.current_contract_label.grid(row=0, column=0, pady=5, sticky='w', padx=10)
        
        self.current_price_label = ttk.Label(current_frame, text="Current: ₹--", font=('Arial', 12))
        self.current_price_label.grid(row=1, column=0, pady=5, sticky='w', padx=10)
        
        self.current_prev_close_label = ttk.Label(current_frame, text="Prev Close: ₹--", font=('Arial', 10))
        self.current_prev_close_label.grid(row=2, column=0, pady=5, sticky='w', padx=10)
        
        self.current_change_label = ttk.Label(current_frame, text="Change: --%", font=('Arial', 10))
        self.current_change_label.grid(row=3, column=0, pady=5, sticky='w', padx=10)
        
        # VS separator
        vs_frame = ttk.Frame(display_frame)
        vs_frame.grid(row=0, column=1, sticky='ns', padx=5, pady=5)
        
        ttk.Label(vs_frame, text="VS", font=('Arial', 16, 'bold')).pack(expand=True)
        
        # Next month frame
        next_frame = ttk.LabelFrame(display_frame, text="Next Month")
        next_frame.grid(row=0, column=2, sticky='nsew', padx=(2, 5), pady=5)
        
        next_frame.grid_columnconfigure(0, weight=1)
        
        self.next_contract_label = ttk.Label(next_frame, text="--", font=('Arial', 10))
        self.next_contract_label.grid(row=0, column=0, pady=5, sticky='w', padx=10)
        
        self.next_price_label = ttk.Label(next_frame, text="Current: ₹--", font=('Arial', 12))
        self.next_price_label.grid(row=1, column=0, pady=5, sticky='w', padx=10)
        
        self.next_prev_close_label = ttk.Label(next_frame, text="Prev Close: ₹--", font=('Arial', 10))
        self.next_prev_close_label.grid(row=2, column=0, pady=5, sticky='w', padx=10)
        
        self.next_change_label = ttk.Label(next_frame, text="Change: --%", font=('Arial', 10))
        self.next_change_label.grid(row=3, column=0, pady=5, sticky='w', padx=10)
        
        # Smiley indicator frame
        smiley_frame = ttk.LabelFrame(display_frame, text="Performance Indicator")
        smiley_frame.grid(row=0, column=3, sticky='nsew', padx=(5, 5), pady=5)
        
        self.month_smiley_label = tk.Label(smiley_frame, text="😐", font=('Arial', 48), bg='white')
        self.month_smiley_label.pack(pady=10)
        
        self.month_comparison_text = ttk.Label(smiley_frame, text="Comparison: --", font=('Arial', 10))
        self.month_comparison_text.pack()
        
        # Middle section: Total Changes Summary
        total_frame = ttk.LabelFrame(right_frame, text="Total Changes Summary")
        total_frame.grid(row=1, column=0, sticky='nsew', pady=5)
        
        total_frame.grid_columnconfigure(1, weight=1)
        
        # Individual changes
        ttk.Label(total_frame, text="Current Month Change:").grid(row=0, column=0, sticky='w', pady=2, padx=10)
        self.total_current_change = ttk.Label(total_frame, text="--%", font=('Arial', 10))
        self.total_current_change.grid(row=0, column=1, sticky='w', pady=2)
        
        ttk.Label(total_frame, text="Next Month Change:").grid(row=1, column=0, sticky='w', pady=2, padx=10)
        self.total_next_change = ttk.Label(total_frame, text="--%", font=('Arial', 10))
        self.total_next_change.grid(row=1, column=1, sticky='w', pady=2)
        
        ttk.Label(total_frame, text="Performance Difference:").grid(row=2, column=0, sticky='w', pady=2, padx=10)
        self.total_perf_diff = ttk.Label(total_frame, text="--%", font=('Arial', 10))
        self.total_perf_diff.grid(row=2, column=1, sticky='w', pady=2)
        
        # TOTAL SUM of changes (NEW FEATURE)
        ttk.Label(total_frame, text="TOTAL SUM of Changes:", 
                 font=('Arial', 11, 'bold')).grid(row=3, column=0, sticky='w', pady=5, padx=10)
        self.total_sum_label = ttk.Label(total_frame, text="--%", 
                                        font=('Arial', 12, 'bold'))
        self.total_sum_label.grid(row=3, column=1, sticky='w', pady=5)
        
        # Price Difference in Rupees section
        price_diff_frame = ttk.LabelFrame(right_frame, text="Price Difference in Rupees")
        price_diff_frame.grid(row=2, column=0, sticky='nsew', pady=5)
        
        price_diff_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Label(price_diff_frame, text="Current Month Change (₹):").grid(row=0, column=0, sticky='w', pady=2, padx=10)
        self.price_diff_current = ttk.Label(price_diff_frame, text="₹--", font=('Arial', 10))
        self.price_diff_current.grid(row=0, column=1, sticky='w', pady=2)
        
        ttk.Label(price_diff_frame, text="Next Month Change (₹):").grid(row=1, column=0, sticky='w', pady=2, padx=10)
        self.price_diff_next = ttk.Label(price_diff_frame, text="₹--", font=('Arial', 10))
        self.price_diff_next.grid(row=1, column=1, sticky='w', pady=2)
        
        ttk.Label(price_diff_frame, text="Price Difference (₹):", 
                 font=('Arial', 11, 'bold')).grid(row=2, column=0, sticky='w', pady=5, padx=10)
        self.price_diff_total = ttk.Label(price_diff_frame, text="₹--", 
                                         font=('Arial', 12, 'bold'))
        self.price_diff_total.grid(row=2, column=1, sticky='w', pady=5)
        
        # Entry/Exit Signal Display
        signal_frame = ttk.LabelFrame(right_frame, text="Entry/Exit Signal")
        signal_frame.grid(row=3, column=0, sticky='nsew', pady=5)
        
        signal_frame.grid_columnconfigure(0, weight=1)
        
        signal_container = ttk.Frame(signal_frame)
        signal_container.pack(fill='both', expand=True,  pady=10)
        
        self.signal_display = tk.Label(signal_container, text="--", 
                                      font=('Arial', 48), bg='white')
        self.signal_display.pack(side='left', padx=20)
        
        self.signal_text = ttk.Label(signal_container, text="No Signal", 
                                    font=('Arial', 12, 'bold'))
        self.signal_text.pack(side='left', padx=20)
        
        # Status frame at bottom
        status_frame = ttk.Frame(right_frame)
        status_frame.grid(row=4, column=0, sticky='nsew', pady=(5, 0))
        
        status_frame.grid_columnconfigure(0, weight=1)
        status_frame.grid_columnconfigure(1, weight=1)
        status_frame.grid_columnconfigure(2, weight=1)
        
        self.trigger_status_label = ttk.Label(status_frame, text="Trigger Status: Ready", foreground='green')
        self.trigger_status_label.grid(row=0, column=0, sticky='w', padx=10, pady=5)
        
        self.month_status_label = ttk.Label(status_frame, text="Status: Not Monitoring", foreground='red')
        self.month_status_label.grid(row=0, column=1, sticky='w', padx=10, pady=5)
        
        self.month_result_label = ttk.Label(status_frame, text="Comparison: --", font=('Arial', 10, 'bold'))
        self.month_result_label.grid(row=0, column=2, sticky='w', padx=10, pady=5)

    def setup_trading_tab(self, notebook):
        """Setup trading tab with buy/sell functionality using resizable frames"""
        trading_frame = ttk.Frame(notebook)
        #notebook.add(trading_frame, text="📈 Trading")
        
        # Configure grid for resizable layout
        trading_frame.grid_columnconfigure(0, weight=1)  # Left column
        trading_frame.grid_columnconfigure(1, weight=1)  # Right column
        trading_frame.grid_rowconfigure(0, weight=1)
        
        # Left frame - Order placement and positions
        left_frame = ttk.Frame(trading_frame)
        left_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 5))
        
        # Right frame - Order history and pending orders
        right_frame = ttk.Frame(trading_frame)
        right_frame.grid(row=0, column=1, sticky='nsew')
        
        # Configure left frame grid
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_rowconfigure(0, weight=1)  # Order placement
        left_frame.grid_rowconfigure(1, weight=1)  # Positions
        
        # Configure right frame grid
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(0, weight=1)  # Order history
        right_frame.grid_rowconfigure(1, weight=1)  # Pending orders
        
        # LEFT FRAME - Order Placement and Current Positions
        # Top section: Order Placement
        order_frame = ttk.LabelFrame(left_frame, text="Order Placement")
        order_frame.grid(row=0, column=0, sticky='nsew', pady=(0, 5))
        
        order_frame.grid_columnconfigure(1, weight=1)
        
        # Contract selection
        ttk.Label(order_frame, text="Contract:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.trading_contract = ttk.Combobox(order_frame, width=25)
        self.trading_contract.grid(row=0, column=1, padx=5, pady=5, sticky='nsew')
        
        # Quantity
        ttk.Label(order_frame, text="Quantity:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        self.quantity_var = tk.StringVar(value="1")
        self.quantity_entry = ttk.Entry(order_frame, textvariable=self.quantity_var, width=10)
        self.quantity_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        # Price type
        ttk.Label(order_frame, text="Price Type:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        self.price_type = ttk.Combobox(order_frame, values=["MARKET", "LIMIT"], width=10)
        self.price_type.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        self.price_type.set("MARKET")
        
        # Limit price (only for LIMIT orders)
        self.limit_price_label = ttk.Label(order_frame, text="Limit Price:")
        self.limit_price_label.grid(row=3, column=0, padx=5, pady=5, sticky='w')
        self.limit_price_var = tk.StringVar()
        self.limit_price_entry = ttk.Entry(order_frame, textvariable=self.limit_price_var, width=10)
        self.limit_price_entry.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        
        # Show/hide limit price based on price type
        self.price_type.bind('<<ComboboxSelected>>', self.toggle_limit_price())
        
        # Product type
        ttk.Label(order_frame, text="Product Type:").grid(row=4, column=0, padx=5, pady=5, sticky='w')
        self.product_type = ttk.Combobox(order_frame, values=["MIS", "CNC", "NRML"], width=10)
        self.product_type.grid(row=4, column=1, padx=5, pady=5, sticky='w')
        self.product_type.set("NRML")
        
        # Action buttons frame
        button_frame = ttk.Frame(order_frame)
        button_frame.grid(row=5, column=0, columnspan=2, pady=15)
        
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)
        button_frame.grid_columnconfigure(2, weight=1)
        button_frame.grid_columnconfigure(3, weight=1)
        
        # Style configuration for buttons
        style = ttk.Style()
        style.configure("Buy.TButton", foreground='green', background='green', font=('Arial', 10, 'bold'))
        style.configure("Sell.TButton", foreground='red', background='red', font=('Arial', 10, 'bold'))
        style.configure("BuyTogether.TButton", foreground='blue', background='blue', font=('Arial', 10, 'bold'))
        style.configure("PlaceOrder.TButton", foreground='purple', background='purple', font=('Arial', 10, 'bold'))
        
        # Single Buy button
        ttk.Button(button_frame, text="BUY", 
                  command=self.place_buy_order, 
                  style="Buy.TButton").grid(row=0, column=0, padx=2, sticky='nsew')
        
        # Single Sell button
        ttk.Button(button_frame, text="SELL", 
                  command=self.place_sell_order, 
                  style="Sell.TButton").grid(row=0, column=1, padx=2, sticky='nsew')
        
        # Buy Together button
        ttk.Button(button_frame, text="BUY TOGETHER", 
                  command=self.place_buy_together_order,
                  style="BuyTogether.TButton").grid(row=0, column=2, padx=2, sticky='nsew')
        
        # Place Order button
        ttk.Button(button_frame, text="PLACE ORDER", 
                  command=self.place_order,
                  style="PlaceOrder.TButton").grid(row=0, column=3, padx=2, sticky='nsew')
        
        # Bottom section: Current Positions
        positions_frame = ttk.LabelFrame(left_frame, text="Current Positions")
        positions_frame.grid(row=1, column=0, sticky='nsew')
        
        positions_frame.grid_columnconfigure(0, weight=1)
        positions_frame.grid_rowconfigure(0, weight=1)
        
        positions_frame.grid_rowconfigure(1, weight=1)
        
        # Positions treeview with scrollbar
        tree_container = ttk.Frame(positions_frame)
        tree_container.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        
        tree_container.grid_columnconfigure(0, weight=1)
        tree_container.grid_rowconfigure(0, weight=1)
        
        columns = ("Contract", "Quantity", "Avg Price", "LTP", "P&L")
        self.positions_tree = ttk.Treeview(tree_container, columns=columns, show="headings", height=8)
        
        for col in columns:
            self.positions_tree.heading(col, text=col)
            self.positions_tree.column(col, width=100)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.positions_tree.yview)
        self.positions_tree.configure(yscrollcommand=scrollbar.set)
        
        self.positions_tree.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')
        
        # Refresh positions button
        ttk.Button(positions_frame, text="Refresh Positions", 
                  command=self.refresh_positions).grid(row=1, column=0, pady=5)
        
        # RIGHT FRAME - Order History and Pending Orders
        # Top section: Order History
        history_frame = ttk.LabelFrame(right_frame, text="Order History")
        history_frame.grid(row=0, column=0, sticky='nsew', pady=(0, 5))
        
        history_frame.grid_columnconfigure(0, weight=1)
        history_frame.grid_rowconfigure(0, weight=1)
        history_frame.grid_rowconfigure(1, weight=1)
        
        # Order history treeview with scrollbar
        history_tree_container = ttk.Frame(history_frame)
        history_tree_container.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        
        history_tree_container.grid_columnconfigure(0, weight=1)
        history_tree_container.grid_rowconfigure(0, weight=1)
        
        history_columns = ("Time", "Contract", "Type", "Qty", "Price", "Status")
        self.order_history_tree = ttk.Treeview(history_tree_container, columns=history_columns, show="headings", height=8)
        
        for col in history_columns:
            self.order_history_tree.heading(col, text=col)
            self.order_history_tree.column(col, width=80)
        
        # Add scrollbar
        history_scrollbar = ttk.Scrollbar(history_tree_container, orient="vertical", command=self.order_history_tree.yview)
        self.order_history_tree.configure(yscrollcommand=history_scrollbar.set)
        
        self.order_history_tree.grid(row=0, column=0, sticky='nsew')
        history_scrollbar.grid(row=0, column=1, sticky='ns')
        
        # Clear history button
        ttk.Button(history_frame, text="Clear History", 
                  command=self.clear_order_history).grid(row=1, column=0, pady=5)
        
        # Bottom section: Pending Orders
        pending_frame = ttk.LabelFrame(right_frame, text="Pending Orders")
        pending_frame.grid(row=1, column=0, sticky='nsew')
        
        pending_frame.grid_columnconfigure(0, weight=1)
        pending_frame.grid_rowconfigure(0, weight=1)
        pending_frame.grid_rowconfigure(1, weight=1)
        
        # Pending orders treeview with scrollbar
        pending_tree_container = ttk.Frame(pending_frame)
        pending_tree_container.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        
        pending_tree_container.grid_columnconfigure(0, weight=1)
        pending_tree_container.grid_rowconfigure(0, weight=1)
        
        pending_columns = ("Time", "Contract", "Type", "Qty", "Price", "Action")
        self.pending_tree = ttk.Treeview(pending_tree_container, columns=pending_columns, show="headings", height=8)
        
        for col in pending_columns:
            self.pending_tree.heading(col, text=col)
            self.pending_tree.column(col, width=80)
        
        # Add scrollbar
        pending_scrollbar = ttk.Scrollbar(pending_tree_container, orient="vertical", command=self.pending_tree.yview)
        self.pending_tree.configure(yscrollcommand=pending_scrollbar.set)
        
        self.pending_tree.grid(row=0, column=0, sticky='nsew')
        pending_scrollbar.grid(row=0, column=1, sticky='ns')
        
        # Cancel order button
        ttk.Button(pending_frame, text="Cancel Selected", 
                  command=self.cancel_selected_order).grid(row=1, column=0, pady=5)
        
        # Auto-populate contracts when month comparison is loaded
        self.trading_contract.bind('<<ComboboxSelected>>', lambda e: self.update_trading_contracts())

    def setup_calendar_spread_tab(self, notebook):
        """Setup calendar spread trading tab with resizable frames"""
        spread_frame = ttk.Frame(notebook)
        notebook.add(spread_frame, text="📊 Calendar Spread")
        
        # Configure grid for resizable layout
        spread_frame.grid_columnconfigure(0, weight=1)  # Left column
        spread_frame.grid_columnconfigure(1, weight=1)  # Right column
        spread_frame.grid_rowconfigure(0, weight=1)
        
        # Left frame - Spread configuration and strategy
        left_frame = ttk.Frame(spread_frame)
        left_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 5))
        
        # Right frame - Spread positions and history
        right_frame = ttk.Frame(spread_frame)
        right_frame.grid(row=0, column=1, sticky='nsew')
        
        # Configure left frame grid
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_rowconfigure(0, weight=1)  # Configuration
        left_frame.grid_rowconfigure(1, weight=1)  # Strategy
        
        # Configure right frame grid
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(0, weight=1)  # Positions
        right_frame.grid_rowconfigure(1, weight=1)  # History
        right_frame.grid_rowconfigure(2, weight=1)  # Performance
        
        # LEFT FRAME - Spread Configuration and Strategy
        # Top section: Spread Configuration
        config_frame = ttk.LabelFrame(left_frame, text="Calendar Spread Configuration")
        config_frame.grid(row=0, column=0, sticky='nsew', pady=(0, 5))
        
        config_frame.grid_columnconfigure(1, weight=1)
        
        # Spread type selection
        ttk.Label(config_frame, text="Spread Type:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.spread_type = ttk.Combobox(config_frame, values=["ENTRY Spread", "EXIT Spread"], width=15)
        self.spread_type.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        self.spread_type.set("ENTRY Spread")
        
        # Spread quantity
        ttk.Label(config_frame, text="Spread Quantity:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        self.spread_quantity_var = tk.StringVar(value="1")
        self.spread_quantity_entry = ttk.Entry(config_frame, textvariable=self.spread_quantity_var, width=10)
        self.spread_quantity_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        
        # Price type for spread
        ttk.Label(config_frame, text="Price Type:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        self.spread_price_type = ttk.Combobox(config_frame, values=["MARKET", "LIMIT"], width=10)
        self.spread_price_type.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        self.spread_price_type.set("MARKET")
        
        # Limit price for spread
        ttk.Label(config_frame, text="Spread Limit Price:").grid(row=3, column=0, padx=5, pady=5, sticky='w')
        self.spread_limit_price_var = tk.StringVar()
        self.spread_limit_price_entry = ttk.Entry(config_frame, textvariable=self.spread_limit_price_var, width=10)
        self.spread_limit_price_entry.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        
        # Product type for spread
        ttk.Label(config_frame, text="Product Type:").grid(row=4, column=0, padx=5, pady=5, sticky='w')
        self.spread_product_type = ttk.Combobox(config_frame, values=["MIS", "CNC", "NRML"], width=10)
        self.spread_product_type.grid(row=4, column=1, padx=5, pady=5, sticky='w')
        self.spread_product_type.set("NRML")
        
        # Auto-trading toggle
        self.auto_spread_var = tk.BooleanVar(value=False)
        self.auto_spread_check = ttk.Checkbutton(config_frame, text="Auto-trade Spreads on Signals", 
                                                variable=self.auto_spread_var)
        self.auto_spread_check.grid(row=5, column=0, columnspan=2, pady=10, sticky='w')
        
        # Spread action buttons
        action_frame = ttk.Frame(config_frame)
        action_frame.grid(row=6, column=0, columnspan=2, pady=15)
        
        action_frame.grid_columnconfigure(0, weight=1)
        action_frame.grid_columnconfigure(1, weight=1)
        action_frame.grid_columnconfigure(2, weight=1)
        action_frame.grid_columnconfigure(3, weight=1)
        
        # Style configuration for spread buttons
        style = ttk.Style()
        style.configure("EntrySpread.TButton", foreground='dark green', background='dark green', font=('Arial', 10, 'bold'))
        style.configure("ExitSpread.TButton", foreground='dark orange', background='dark orange', font=('Arial', 10, 'bold'))
        style.configure("CloseSpread.TButton", foreground='purple', background='purple', font=('Arial', 10, 'bold'))
        style.configure("TestSpread.TButton", foreground='black', background='black', font=('Arial', 10, 'bold'))
        
        # Place Entry Spread button
        ttk.Button(action_frame, text="Place ENTRY Spread", 
                  command=lambda: self.place_calendar_spread("ENTRY"),
                  style="EntrySpread.TButton").grid(row=0, column=0, padx=2, sticky='nsew')
        
        # Place Exit Spread button
        ttk.Button(action_frame, text="Place EXIT Spread", 
                  command=lambda: self.place_calendar_spread("EXIT"),
                  style="ExitSpread.TButton").grid(row=0, column=1, padx=2, sticky='nsew')
        
        # Close Spread button
        ttk.Button(action_frame, text="Close Spread", 
                  command=self.close_calendar_spread,
                  style="CloseSpread.TButton").grid(row=0, column=2, padx=2, sticky='nsew')
        
        # Test Spread button
        ttk.Button(action_frame, text="Test Spread", 
                  command=self.test_spread_order, style="TestSpread.TButton").grid(row=0, column=3, padx=2, sticky='nsew')
        
        # Bottom section: Spread Strategy Explanation
        strategy_frame = ttk.LabelFrame(left_frame, text="Spread Strategy Logic")
        strategy_frame.grid(row=1, column=0, sticky='nsew')
        
        strategy_frame.grid_columnconfigure(0, weight=1)
        strategy_frame.grid_rowconfigure(0, weight=1)
        
        strategy_text = """
        📈 ENTRY SPREAD STRATEGY:
        When price difference < entry threshold (-6₹):
        1. BUY Next Month (cheaper/outperforming)
        2. SELL Current Month (expensive/underperforming)
        → Betting on convergence
        
        📉 EXIT SPREAD STRATEGY:
        When price difference > exit threshold (+6₹):
        1. BUY Current Month (undervalued)
        2. SELL Next Month (overvalued)
        → Betting on mean reversion
        
        💡 The spread profits from the price difference
        between months narrowing or widening.
        """
        strategy_label = ttk.Label(strategy_frame, text=strategy_text, justify='left', 
                                  font=('Arial', 9), wraplength=300)
        strategy_label.grid(row=0, column=0, sticky='nw', padx=10, pady=10)
        
        # RIGHT FRAME - Spread Positions and History
        # Top section: Current Spread Positions
        spread_positions_frame = ttk.LabelFrame(right_frame, text="Current Spread Positions")
        spread_positions_frame.grid(row=0, column=0, sticky='nsew', pady=(0, 5))
        
        spread_positions_frame.grid_columnconfigure(0, weight=1)
        spread_positions_frame.grid_rowconfigure(0, weight=1)
        spread_positions_frame.grid_rowconfigure(1, weight=1)
        
        # Spread positions treeview with scrollbar
        spread_positions_tree_container = ttk.Frame(spread_positions_frame)
        spread_positions_tree_container.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        
        spread_positions_tree_container.grid_columnconfigure(0, weight=1)
        spread_positions_tree_container.grid_rowconfigure(0, weight=1)
        
        spread_columns = ("Leg", "Contract", "Type", "Qty", "Avg Price", "P&L")
        self.spread_positions_tree = ttk.Treeview(spread_positions_tree_container, columns=spread_columns, show="headings", height=6)
        
        for col in spread_columns:
            self.spread_positions_tree.heading(col, text=col)
            self.spread_positions_tree.column(col, width=80)
        
        # Add scrollbar
        spread_scrollbar = ttk.Scrollbar(spread_positions_tree_container, orient="vertical", command=self.spread_positions_tree.yview)
        self.spread_positions_tree.configure(yscrollcommand=spread_scrollbar.set)
        
        self.spread_positions_tree.grid(row=0, column=0, sticky='nsew')
        spread_scrollbar.grid(row=0, column=1, sticky='ns')
        
        # Refresh spread positions button
        ttk.Button(spread_positions_frame, text="Refresh Spreads", 
                  command=self.refresh_spread_positions).grid(row=1, column=0, pady=5)
        
        # Middle section: Spread Order History
        spread_history_frame = ttk.LabelFrame(right_frame, text="Spread Order History")
        spread_history_frame.grid(row=1, column=0, sticky='nsew', pady=5)
        
        spread_history_frame.grid_columnconfigure(0, weight=1)
        spread_history_frame.grid_rowconfigure(0, weight=1)
        spread_history_frame.grid_rowconfigure(1, weight=1)
        
        # Spread history treeview with scrollbar
        spread_history_tree_container = ttk.Frame(spread_history_frame)
        spread_history_tree_container.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        
        spread_history_tree_container.grid_columnconfigure(0, weight=1)
        spread_history_tree_container.grid_rowconfigure(0, weight=1)
        
        spread_history_columns = ("Time", "Type", "Leg1", "Leg2", "Qty", "Status")
        self.spread_history_tree = ttk.Treeview(spread_history_tree_container, columns=spread_history_columns, show="headings", height=6)
        
        for col in spread_history_columns:
            self.spread_history_tree.heading(col, text=col)
            self.spread_history_tree.column(col, width=80)
        
        # Add scrollbar
        spread_history_scrollbar = ttk.Scrollbar(spread_history_tree_container, orient="vertical", command=self.spread_history_tree.yview)
        self.spread_history_tree.configure(yscrollcommand=spread_history_scrollbar.set)
        
        self.spread_history_tree.grid(row=0, column=0, sticky='nsew')
        spread_history_scrollbar.grid(row=0, column=1, sticky='ns')
        
        # Clear spread history button
        ttk.Button(spread_history_frame, text="Clear History", 
                  command=self.clear_spread_history).grid(row=1, column=0, pady=5)
        
        # Bottom section: Spread Performance Summary
        performance_frame = ttk.LabelFrame(right_frame, text="Spread Performance Summary")
        performance_frame.grid(row=2, column=0, sticky='nsew')
        
        performance_frame.grid_columnconfigure(1, weight=1)
        
        # Performance metrics
        ttk.Label(performance_frame, text="Total Spreads Placed:").grid(row=0, column=0, sticky='w', pady=2, padx=10)
        self.total_spreads_label = ttk.Label(performance_frame, text="0", font=('Arial', 10, 'bold'))
        self.total_spreads_label.grid(row=0, column=1, sticky='w', pady=2)
        
        ttk.Label(performance_frame, text="Successful Spreads:").grid(row=1, column=0, sticky='w', pady=2, padx=10)
        self.successful_spreads_label = ttk.Label(performance_frame, text="0", font=('Arial', 10, 'bold'), foreground='green')
        self.successful_spreads_label.grid(row=1, column=1, sticky='w', pady=2)
        
        ttk.Label(performance_frame, text="Failed Spreads:").grid(row=2, column=0, sticky='w', pady=2, padx=10)
        self.failed_spreads_label = ttk.Label(performance_frame, text="0", font=('Arial', 10, 'bold'), foreground='red')
        self.failed_spreads_label.grid(row=2, column=1, sticky='w', pady=2)
        
        ttk.Label(performance_frame, text="Net Spread P&L:").grid(row=3, column=0, sticky='w', pady=5, padx=10)
        self.net_spread_pnl_label = ttk.Label(performance_frame, text="₹0.00", font=('Arial', 12, 'bold'))
        self.net_spread_pnl_label.grid(row=3, column=1, sticky='w', pady=5)

    def setup_auto_exit_tab(self, notebook):
        """Setup auto-exit tab for profit monitoring"""
        auto_exit_frame = ttk.Frame(notebook)
        notebook.add(auto_exit_frame, text="🤖 Auto Exit")
        
        # Configure grid for resizable layout
        auto_exit_frame.grid_columnconfigure(0, weight=1)
        auto_exit_frame.grid_rowconfigure(0, weight=1)
        
        # Main container
        container = ttk.Frame(auto_exit_frame)
        container.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Auto-exit settings frame
        settings_frame = ttk.LabelFrame(container, text="Auto Exit Settings")
        settings_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        settings_frame.grid_columnconfigure(1, weight=1)
        
        # Profit target
        ttk.Label(settings_frame, text="Profit Target (₹):").grid(row=0, column=0, padx=10, pady=10, sticky='w')
        self.profit_target_var = tk.StringVar(value="10000000.0")
        self.profit_target_entry = ttk.Entry(settings_frame, textvariable=self.profit_target_var, width=15)
        self.profit_target_entry.grid(row=0, column=1, padx=10, pady=10, sticky='w')
        
        # Stop loss
        if 0:
            ttk.Label(settings_frame, text="Stop Loss (₹):").grid(row=1, column=0, padx=10, pady=10, sticky='w')
            self.stop_loss_var = tk.StringVar(value="50.0")
            self.stop_loss_entry = ttk.Entry(settings_frame, textvariable=self.stop_loss_var, width=15)
            self.stop_loss_entry.grid(row=1, column=1, padx=10, pady=10, sticky='w')
        
        # Checkbox for auto-exit
        self.auto_exit_var = tk.BooleanVar(value=False)
        auto_exit_check = ttk.Checkbutton(settings_frame, text="Enable Auto Exit", 
                                         variable=self.auto_exit_var,
                                         command=self.toggle_auto_exit)
        auto_exit_check.grid(row=2, column=0, columnspan=2, pady=10, sticky='w')
        
        # Control buttons
        button_frame = ttk.Frame(settings_frame)
        button_frame.grid(row=3, column=0, columnspan=2, pady=15)
        
        ttk.Button(button_frame, text="Start Monitoring", 
                  command=self.start_auto_exit_monitoring).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Stop Monitoring", 
                  command=self.stop_auto_exit_monitoring).pack(side='left', padx=5)
        
        # Status frame
        status_frame = ttk.LabelFrame(container, text="Auto Exit Status")
        status_frame.pack(fill='both', expand=True, pady=10)
        
        # Current position info
        ttk.Label(status_frame, text="Current Position:").grid(row=0, column=0, padx=10, pady=10, sticky='w')
        self.position_status_label = ttk.Label(status_frame, text="No active position", foreground='red')
        self.position_status_label.grid(row=0, column=1, padx=10, pady=10, sticky='w')
        
        # Entry price
        ttk.Label(status_frame, text="Entry Price:").grid(row=1, column=0, padx=10, pady=5, sticky='w')
        self.entry_price_label = ttk.Label(status_frame, text="₹0.00")
        self.entry_price_label.grid(row=1, column=1, padx=10, pady=5, sticky='w')
        
        # Current P&L
        ttk.Label(status_frame, text="Current P&L:").grid(row=2, column=0, padx=10, pady=5, sticky='w')
        self.current_pnl_label = ttk.Label(status_frame, text="₹0.00", font=('Arial', 12, 'bold'))
        self.current_pnl_label.grid(row=2, column=1, padx=10, pady=5, sticky='w')
        
        # Profit target status
        ttk.Label(status_frame, text="Profit Target:").grid(row=3, column=0, padx=10, pady=5, sticky='w')
        self.profit_target_status = ttk.Label(status_frame, text="₹10000000.00")
        self.profit_target_status.grid(row=3, column=1, padx=10, pady=5, sticky='w')
        
        # Stop loss status
        # ttk.Label(status_frame, text="Stop Loss:").grid(row=4, column=0, padx=10, pady=5, sticky='w')
        # self.stop_loss_status = ttk.Label(status_frame, text="₹50.00")
        # self.stop_loss_status.grid(row=4, column=1, padx=10, pady=5, sticky='w')
        
        # Action button
        self.exit_button = ttk.Button(status_frame, text="MANUAL EXIT", 
                                     command=self.manual_exit_position,
                                     state='disabled')
        self.exit_button.grid(row=5, column=0, columnspan=2, pady=15)
        
        # Log frame
        log_frame = ttk.LabelFrame(container, text="Auto Exit Log")
        log_frame.pack(fill='both', expand=True, pady=10)
        
        log_frame.grid_columnconfigure(0, weight=1)
        log_frame.grid_rowconfigure(0, weight=1)
        
        self.auto_exit_log = scrolledtext.ScrolledText(log_frame, height=10)
        self.auto_exit_log.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)

    def toggle_limit_price(self, event=None):
        """Show/hide limit price field based on price type"""
        if self.price_type.get() == "LIMIT":
            self.limit_price_label.grid()
            self.limit_price_entry.grid()
        else:
            self.limit_price_label.grid_remove()
            self.limit_price_entry.grid_remove()

    def get_contracts_list(self):
        return self.current_month_contract, self.next_month_contract
        
    def update_trading_contracts(self):
        try:
            if hasattr(self, 'current_month_contract') and hasattr(self, 'next_month_contract'):
                contracts = self.get_contracts_list()
                
                if not contracts:
                    print("Warning: No contracts available")
                    self.trading_contract.set('')
                    self.trading_contract['values'] = []
                    return
                
                # Ensure we have valid string values
                valid_contracts = [str(c) for c in contracts if c]
                
                if not valid_contracts:
                    print("Warning: No contracts available")
                    self.trading_contract.set('')
                    self.trading_contract['values'] = []
                    return
                    
                # Update combobox values
                self.trading_contract['values'] = valid_contracts
                
                # Set the first contract
                self.trading_contract.set(valid_contracts[0])
                
        except Exception as e:
            print(f"Error updating contracts: {e}")
            self.trading_contract.set('')

    def load_month_contracts(self):
        """Load current and next month contracts"""
        if not self.is_logged_in:
            messagebox.showerror("Error", "Please login first")
            return
        
        commodity = self.month_commodity.get()
        
        try:
            contracts = self.get_monthly_contracts(commodity)
            
            if len(contracts) < 2:
                messagebox.showerror("Error", f"Need at least 2 contracts for {commodity}")
                return
            
            # Store current and next month contracts
            self.current_month_contract = contracts[0]
            self.next_month_contract = contracts[1]
            
            # Update contract labels
            self.current_contract_label.config(text=self.current_month_contract)
            self.next_contract_label.config(text=self.next_month_contract)
            
            # Update trading tab contracts
            self.update_trading_contracts()
            
            self.log_message(f"Loaded month comparison: {self.current_month_contract} vs {self.next_month_contract}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load contracts: {e}")

    def place_buy_order(self):
        """Place a single BUY order"""
        self.place_order("BUY")

    def place_sell_order(self):
        """Place a single SELL order"""
        self.place_order("SELL")

    def place_buy_together_order(self):
        """Place BUY orders for both current and next month contracts"""
        if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
            messagebox.showerror("Error", "Please load contracts first")
            return
        
        try:
            quantity = int(self.quantity_var.get())
            price_type = self.price_type.get()
            product = self.product_type.get()
            
            # Place order for current month
            current_result = self.execute_order(
                tradingsymbol=self.current_month_contract,
                transaction_type="BUY",
                quantity=quantity,
                order_type=price_type,
                product=product,
                price=float(self.limit_price_var.get()) if price_type == "LIMIT" and self.limit_price_var.get() else None
            )
            
            # Place order for next month
            next_result = self.execute_order(
                tradingsymbol=self.next_month_contract,
                transaction_type="BUY",
                quantity=quantity,
                order_type=price_type,
                product=product,
                price=float(self.limit_price_var.get()) if price_type == "LIMIT" and self.limit_price_var.get() else None
            )
            
            if current_result and next_result:
                messagebox.showinfo("Success", "BUY TOGETHER orders placed successfully!")
                
                # NEW: Check for auto-exit monitoring
                self.root.after(2000, self.check_current_position)
                if self.auto_exit_running and self.current_position:
                    self.log_auto_exit(f"New BUY TOGETHER position detected")
            else:
                messagebox.showwarning("Partial Success", 
                                     "Some orders may not have been placed. Check order history.")
                
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to place orders: {e}")

    def place_order(self, transaction_type=None):
        """Place an order with current settings"""
        if not self.is_logged_in:
            messagebox.showerror("Error", "Please login first")
            return
        
        contract = self.trading_contract.get()
        if not contract:
            messagebox.showerror("Error", "Please select a contract")
            return
        
        try:
            if not transaction_type:
                # This is for the generic PLACE ORDER button - we need to ask for transaction type
                transaction_type = self.ask_transaction_type()
                if not transaction_type:
                    return
            
            quantity = int(self.quantity_var.get())
            price_type = self.price_type.get()
            product = self.product_type.get()
            
            price = None
            if price_type == "LIMIT":
                if not self.limit_price_var.get():
                    messagebox.showerror("Error", "Please enter limit price")
                    return
                price = float(self.limit_price_var.get())
            
            result = self.execute_order(
                tradingsymbol=contract,
                transaction_type=transaction_type,
                quantity=quantity,
                order_type=price_type,
                product=product,
                price=price
            )
            
            if result:
                message = f"{transaction_type} order placed successfully!"
                messagebox.showinfo("Success", message)
                
                # NEW: Check for auto-exit monitoring if position is opened
                if transaction_type in ["BUY", "SELL"]:
                    # Wait a moment for position to update
                    self.root.after(2000, self.check_current_position)
                    if self.auto_exit_running and self.current_position:
                        self.log_auto_exit(f"New position detected: {contract} {transaction_type} {quantity}")
            else:
                messagebox.showerror("Error", "Failed to place order")
                
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to place order: {e}")

    def ask_transaction_type(self):
        """Ask user for transaction type (BUY/SELL)"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Select Transaction Type")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Make dialog resizable
        dialog.resizable(True, True)
        
        ttk.Label(dialog, text="Select Transaction Type:", 
                 font=('Arial', 10, 'bold')).pack(pady=20)
        
        result = {"type": None}
        
        def select_buy():
            result["type"] = "BUY"
            dialog.destroy()
        
        def select_sell():
            result["type"] = "SELL"
            dialog.destroy()
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="BUY", command=select_buy, 
                  style="Buy.TButton").pack(side='left', padx=10)
        ttk.Button(button_frame, text="SELL", command=select_sell,
                  style="Sell.TButton").pack(side='left', padx=10)
        
        dialog.wait_window()
        return result["type"]

    def execute_order(self, tradingsymbol, transaction_type, quantity, order_type, product, price=None):
        """Execute an order through Zerodha API"""
        try:
            # Validate inputs
            if quantity <= 0:
                raise ValueError("Quantity must be positive")
            
            # Prepare order parameters
            order_params = {
                "tradingsymbol": tradingsymbol,
                "exchange": "MCX",
                "transaction_type": transaction_type,
                "quantity": quantity,
                "order_type": order_type,
                "product": product,
                "validity": "DAY"
            }
            
            # Add price for limit orders
            if order_type == "LIMIT" and price:
                order_params["price"] = price
                order_params["validity"] = "DAY"
            
            # Place the order
            order_id = self.kite.place_order(
                variety="regular",
                **order_params
            )
            
            # Log the order
            order_time = datetime.now().strftime("%H:%M:%S")
            order_price = price if price else "MARKET"
            
            order_record = {
                "time": order_time,
                "contract": tradingsymbol,
                "type": transaction_type,
                "quantity": quantity,
                "price": order_price,
                "status": "PLACED",
                "order_id": order_id
            }
            
            # Add to order history
            self.order_history.append(order_record)
            
            # Add to pending orders
            self.pending_orders.append(order_record)
            
            # Update UI
            self.update_order_history_display()
            self.update_pending_orders_display()
            
            # Log message
            self.log_message(f"✅ Order placed: {transaction_type} {quantity} {tradingsymbol} at {order_price}")
            
            return True
            
        except Exception as e:
            self.log_message(f"❌ Order failed: {transaction_type} {quantity} {tradingsymbol} - {str(e)}")
            return False

    def update_order_history_display(self):
        """Update the order history treeview"""
        # Clear existing items
        for item in self.order_history_tree.get_children():
            self.order_history_tree.delete(item)
        
        # Add order history items
        for order in self.order_history[-50:]:  # Show last 50 orders
            values = (
                order["time"],
                order["contract"],
                order["type"],
                order["quantity"],
                order["price"],
                order["status"]
            )
            
            item = self.order_history_tree.insert("", "end", values=values)
            
            # Color code based on order type
            if order["type"] == "BUY":
                self.order_history_tree.item(item, tags=("buy",))
            else:
                self.order_history_tree.item(item, tags=("sell",))
        
        # Configure tags
        self.order_history_tree.tag_configure("buy", foreground="green")
        self.order_history_tree.tag_configure("sell", foreground="red")

    def update_pending_orders_display(self):
        """Update the pending orders treeview"""
        # Clear existing items
        for item in self.pending_tree.get_children():
            self.pending_tree.delete(item)
        
        # Add pending order items
        for order in self.pending_orders:
            values = (
                order["time"],
                order["contract"],
                order["type"],
                order["quantity"],
                order["price"],
                "Cancel"
            )
            
            item = self.pending_tree.insert("", "end", values=values)
            
            # Store order_id in item
            self.pending_tree.set(item, "order_id", order.get("order_id", ""))
            
            # Color code based on order type
            if order["type"] == "BUY":
                self.pending_tree.item(item, tags=("buy",))
            else:
                self.pending_tree.item(item, tags=("sell",))
        
        # Configure tags
        self.pending_tree.tag_configure("buy", foreground="green")
        self.pending_tree.tag_configure("sell", foreground="red")

    def refresh_positions(self):
        """Refresh current positions from Zerodha"""
        if not self.is_logged_in:
            messagebox.showerror("Error", "Please login first")
            return
        
        try:
            # Get positions from Zerodha
            positions_data = self.kite.positions()
            
            # Clear existing items
            for item in self.positions_tree.get_children():
                self.positions_tree.delete(item)
            
            # Parse positions
            day_positions = positions_data.get('day', [])
            net_positions = positions_data.get('net', [])
            
            all_positions = day_positions + net_positions
            
            for position in all_positions:
                if position['exchange'] == 'MCX' and position['quantity'] != 0:
                    tradingsymbol = position['tradingsymbol']
                    quantity = position['quantity']
                    avg_price = position['average_price']
                    
                    # Get last traded price
                    try:
                        ltp = self.kite.ltp(f"MCX:{tradingsymbol}")[f"MCX:{tradingsymbol}"]['last_price']
                    except:
                        ltp = 0
                    
                    # Calculate P&L
                    if quantity > 0:  # Long position
                        pnl = (ltp - avg_price) * abs(quantity)
                    else:  # Short position
                        pnl = (avg_price - ltp) * abs(quantity)
                    
                    values = (
                        tradingsymbol,
                        quantity,
                        f"₹{avg_price:.2f}",
                        f"₹{ltp:.2f}",
                        f"₹{pnl:+.2f}"
                    )
                    
                    item = self.positions_tree.insert("", "end", values=values)
                    
                    # Color code P&L
                    if pnl > 0:
                        self.positions_tree.item(item, tags=("profit",))
                    elif pnl < 0:
                        self.positions_tree.item(item, tags=("loss",))
            
            # Configure tags
            self.positions_tree.tag_configure("profit", foreground="green")
            self.positions_tree.tag_configure("loss", foreground="red")
            
            self.log_message("Positions refreshed successfully")
            
        except Exception as e:
            self.log_message(f"Error refreshing positions: {e}")

    def cancel_selected_order(self):
        """Cancel the selected pending order"""
        selection = self.pending_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an order to cancel")
            return
        
        item = selection[0]
        order_id = self.pending_tree.set(item, "order_id")
        
        if not order_id:
            messagebox.showerror("Error", "No order ID found for selected item")
            return
        
        try:
            # Cancel the order
            self.kite.cancel_order(
                variety="regular",
                order_id=order_id
            )
            
            # Update order status in history
            for order in self.order_history:
                if order.get("order_id") == order_id:
                    order["status"] = "CANCELLED"
                    break
            
            # Remove from pending orders
            self.pending_orders = [o for o in self.pending_orders if o.get("order_id") != order_id]
            
            # Update displays
            self.update_order_history_display()
            self.update_pending_orders_display()
            
            messagebox.showinfo("Success", "Order cancelled successfully")
            self.log_message(f"Order {order_id} cancelled")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to cancel order: {e}")

    def clear_order_history(self):
        """Clear the order history"""
        if messagebox.askyesno("Confirm", "Are you sure you want to clear order history?"):
            self.order_history = []
            self.update_order_history_display()
            self.log_message("Order history cleared")

    # ==============================================
    # CALENDAR SPREAD TRADING FUNCTIONS
    # ==============================================

    def place_calendar_spread(self, spread_type="ENTRY"):
        """Place a calendar spread order based on signal type"""
        if not self.is_logged_in:
            messagebox.showerror("Error", "Please login first")
            return
        
        if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
            messagebox.showerror("Error", "Please load contracts first")
            return
        
        try:
            # Get spread parameters
            quantity = int(self.spread_quantity_var.get())
            price_type = self.spread_price_type.get()
            product = self.spread_product_type.get()
            limit_price = None
            
            if price_type == "LIMIT":
                if not self.spread_limit_price_var.get():
                    messagebox.showerror("Error", "Please enter spread limit price")
                    return
                limit_price = float(self.spread_limit_price_var.get())
            
            # Define spread legs based on spread type
            if spread_type == "ENTRY":
                # ENTRY Spread: BUY Next Month, SELL Current Month
                leg1 = {
                    "tradingsymbol": self.next_month_contract,
                    "transaction_type": "BUY",
                    "quantity": quantity
                }
                leg2 = {
                    "tradingsymbol": self.current_month_contract,
                    "transaction_type": "SELL",
                    "quantity": quantity
                }
                spread_name = "ENTRY Spread"
            else:  # EXIT Spread
                # EXIT Spread: BUY Current Month, SELL Next Month
                leg1 = {
                    "tradingsymbol": self.current_month_contract,
                    "transaction_type": "BUY",
                    "quantity": quantity
                }
                leg2 = {
                    "tradingsymbol": self.next_month_contract,
                    "transaction_type": "SELL",
                    "quantity": quantity
                }
                spread_name = "EXIT Spread"
            
            # Execute spread orders
            success = self.execute_spread_order(leg1, leg2, price_type, product, limit_price, spread_name)
            
            if success:
                messagebox.showinfo("Success", f"{spread_name} placed successfully!")
                self.log_message(f"✅ {spread_name} placed: {quantity} lots")
            else:
                messagebox.showwarning("Partial Success", 
                                     f"{spread_name} may not have been fully placed. Check spread history.")
                
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to place spread: {e}")

    def execute_spread_order(self, leg1, leg2, order_type, product, price=None, spread_name=""):
        """Execute a calendar spread (two legs simultaneously)"""
        try:
            spread_time = datetime.now().strftime("%H:%M:%S")
            spread_id = f"SPREAD_{int(time.time())}"
            
            # Execute first leg
            leg1_result = self.execute_order(
                tradingsymbol=leg1["tradingsymbol"],
                transaction_type=leg1["transaction_type"],
                quantity=leg1["quantity"],
                order_type=order_type,
                product=product,
                price=price
            )
            
            # Execute second leg
            leg2_result = self.execute_order(
                tradingsymbol=leg2["tradingsymbol"],
                transaction_type=leg2["transaction_type"],
                quantity=leg2["quantity"],
                order_type=order_type,
                product=product,
                price=price
            )
            
            # Record spread order
            spread_record = {
                "time": spread_time,
                "spread_id": spread_id,
                "spread_type": spread_name,
                "leg1": leg1,
                "leg2": leg2,
                "order_type": order_type,
                "product": product,
                "price": price if price else "MARKET",
                "status": "PLACED" if leg1_result and leg2_result else "PARTIAL",
                "success": leg1_result and leg2_result
            }
            
            # Add to spread orders
            self.spread_orders.append(spread_record)
            
            # Update spread position
            self.update_spread_position(spread_record)
            
            # Update UI
            self.update_spread_history_display()
            self.update_spread_performance_metrics()
            
            return leg1_result and leg2_result
            
        except Exception as e:
            self.log_message(f"❌ Spread order failed: {spread_name} - {str(e)}")
            return False

    def update_spread_position(self, spread_record):
        """Update current spread position"""
        try:
            # For simplicity, track the latest spread
            # In a real system, you'd track multiple spreads and their P&L
            self.spread_position = {
                "time": spread_record["time"],
                "spread_type": spread_record["spread_type"],
                "leg1": spread_record["leg1"],
                "leg2": spread_record["leg2"],
                "quantity": spread_record["leg1"]["quantity"],
                "entry_price_diff": self.get_current_price_difference()  # Store entry price difference
            }
            
            self.log_message(f"📊 Spread position updated: {spread_record['spread_type']}")
            
        except Exception as e:
            self.log_message(f"Error updating spread position: {e}")

    def close_calendar_spread(self):
        """Close the current calendar spread position"""
        if not self.spread_position:
            messagebox.showinfo("No Spread", "No active spread position to close")
            return
        
        try:
            spread_type = self.spread_position["spread_type"]
            quantity = self.spread_position["quantity"]
            
            # Determine closing legs (opposite of opening legs)
            if "ENTRY" in spread_type:
                # Close ENTRY Spread: SELL Next Month, BUY Current Month
                leg1 = {
                    "tradingsymbol": self.next_month_contract,
                    "transaction_type": "SELL",  # Opposite of opening BUY
                    "quantity": quantity
                }
                leg2 = {
                    "tradingsymbol": self.current_month_contract,
                    "transaction_type": "BUY",  # Opposite of opening SELL
                    "quantity": quantity
                }
                close_name = "Close ENTRY Spread"
            else:  # EXIT Spread
                # Close EXIT Spread: SELL Current Month, BUY Next Month
                leg1 = {
                    "tradingsymbol": self.current_month_contract,
                    "transaction_type": "SELL",  # Opposite of opening BUY
                    "quantity": quantity
                }
                leg2 = {
                    "tradingsymbol": self.next_month_contract,
                    "transaction_type": "BUY",  # Opposite of opening SELL
                    "quantity": quantity
                }
                close_name = "Close EXIT Spread"
            
            # Get current price type and product
            price_type = self.spread_price_type.get()
            product = self.spread_product_type.get()
            limit_price = None
            
            if price_type == "LIMIT" and self.spread_limit_price_var.get():
                limit_price = float(self.spread_limit_price_var.get())
            
            # Execute closing spread
            success = self.execute_spread_order(leg1, leg2, price_type, product, limit_price, close_name)
            
            if success:
                # Calculate P&L for the closed spread
                entry_diff = self.spread_position.get("entry_price_diff", 0)
                current_diff = self.get_current_price_difference()
                
                # P&L logic depends on spread type
                if "ENTRY" in spread_type:
                    # ENTRY spread profits when price difference increases (becomes less negative or positive)
                    pnl = (current_diff - entry_diff) * quantity
                else:  # EXIT spread
                    # EXIT spread profits when price difference decreases (becomes less positive or negative)
                    pnl = (entry_diff - current_diff) * quantity
                
                # Record closed spread
                closed_record = {
                    "time": datetime.now().strftime("%H:%M:%S"),
                    "spread_type": f"Closed {spread_type}",
                    "entry_diff": entry_diff,
                    "exit_diff": current_diff,
                    "pnl": pnl,
                    "quantity": quantity
                }
                
                self.log_message(f"💰 Spread closed: P&L = ₹{pnl:+.2f}")
                
                # Clear spread position
                self.spread_position = None
                
                messagebox.showinfo("Success", f"{close_name} executed successfully!\nP&L: ₹{pnl:+.2f}")
            else:
                messagebox.showwarning("Partial Close", "Spread may not have been fully closed")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to close spread: {e}")

    def test_entry_exit_popup(self):
        """Test the entry/exit popup display"""
        if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
            messagebox.showerror("Error", "Please load contracts first")
            return
        
        # Test entry popup
        self.show_entry_exit_popup(-10.0, "ENTRY")
        
        # Test exit popup after 2 seconds
        self.root.after(2000, lambda: self.show_entry_exit_popup(2.5, "EXIT"))

    def check_entry_exit_condition(self, price_difference):
        """
        Check if price difference triggers entry or exit condition
        Returns: (should_trigger, signal_type, price_difference)
        """
        try:
            # Update thresholds from GUI
            self.entry_threshold = float(self.entry_threshold_var.get())
            self.exit_threshold = float(self.exit_threshold_var.get())
            self.entry_exit_cooldown = int(self.entry_exit_cooldown_var.get()) * 60  # Convert to seconds
            
            # Check cooldown
            current_time = time.time()
            if self.last_entry_exit_trigger_time is not None and \
               (current_time - self.last_entry_exit_trigger_time) < self.entry_exit_cooldown:
                return False, None, price_difference
            
            # Check conditions
            if price_difference < self.entry_threshold:
                return True, "ENTRY", price_difference
            elif price_difference > self.exit_threshold:
                return True, "EXIT", price_difference
            
            return False, None, price_difference
            
        except ValueError:
            # If invalid thresholds, use defaults
            if price_difference < -10.0:
                return True, "ENTRY", price_difference
            elif price_difference > 10.0:
                return True, "EXIT", price_difference
            return False, None, price_difference
        
    def get_current_price_difference(self):
        """Get current price difference between months"""
        try:
            if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
                return 0
            
            # Get current prices
            contracts = [self.current_month_contract, self.next_month_contract]
            instruments = [f"MCX:{contract}" for contract in contracts]
            quote_data = self.kite.quote(instruments)
            
            current_price = quote_data[f"MCX:{self.current_month_contract}"]['last_price']
            next_price = quote_data[f"MCX:{self.next_month_contract}"]['last_price']
            
            # Get PREVIOUS DAY CLOSE prices
            current_prev = self.previous_day_close_prices.get(self.current_month_contract, current_price)
            next_prev = self.previous_day_close_prices.get(self.next_month_contract, next_price)
            
            # Calculate price difference using the formula:
            # change difference in rs = current month (Current Price - Previous Close) - next month (Current Price - Previous Close)
            current_change_rupees = current_price - current_prev
            next_change_rupees = next_price - next_prev
            price_difference = current_change_rupees - next_change_rupees
            
            return price_difference
            
        except Exception as e:
            self.log_message(f"Error getting current price difference: {e}")
            return 0

    def refresh_spread_positions(self):
        """Refresh current spread positions display"""
        try:
            # Clear existing items
            for item in self.spread_positions_tree.get_children():
                self.spread_positions_tree.delete(item)
            
            # Get current positions from Zerodha
            positions_data = self.kite.positions()
            day_positions = positions_data.get('day', [])
            net_positions = positions_data.get('net', [])
            all_positions = day_positions + net_positions
            
            # Filter for MCX futures positions
            futures_positions = []
            for position in all_positions:
                if position['exchange'] == 'MCX' and position['instrument_type'] == 'FUT':
                    futures_positions.append(position)
            
            # Identify potential spread positions
            # This is a simplified approach - in reality you'd track spread IDs
            if len(futures_positions) >= 2:
                # Sort by tradingsymbol
                futures_positions.sort(key=lambda x: x['tradingsymbol'])
                
                # Display as spread legs
                for i, position in enumerate(futures_positions[:2], 1):
                    tradingsymbol = position['tradingsymbol']
                    quantity = position['quantity']
                    avg_price = position['average_price']
                    
                    # Get last traded price
                    try:
                        ltp = self.kite.ltp(f"MCX:{tradingsymbol}")[f"MCX:{tradingsymbol}"]['last_price']
                    except:
                        ltp = 0
                    
                    # Calculate P&L
                    if quantity > 0:  # Long position
                        pnl = (ltp - avg_price) * abs(quantity)
                        position_type = "LONG"
                    else:  # Short position
                        pnl = (avg_price - ltp) * abs(quantity)
                        position_type = "SHORT"
                    
                    values = (
                        f"Leg {i}",
                        tradingsymbol,
                        position_type,
                        quantity,
                        f"₹{avg_price:.2f}",
                        f"₹{pnl:+.2f}"
                    )
                    
                    item = self.spread_positions_tree.insert("", "end", values=values)
                    
                    # Color code P&L
                    if pnl > 0:
                        self.spread_positions_tree.item(item, tags=("profit",))
                    elif pnl < 0:
                        self.spread_positions_tree.item(item, tags=("loss",))
            
            # Configure tags
            self.spread_positions_tree.tag_configure("profit", foreground="green")
            self.spread_positions_tree.tag_configure("loss", foreground="red")
            
            # Also update if we have a tracked spread position
            if self.spread_position:
                # Add tracked spread position
                values = (
                    "Tracked",
                    self.spread_position["spread_type"],
                    "ACTIVE",
                    self.spread_position["quantity"],
                    f"Entry Diff: ₹{self.spread_position.get('entry_price_diff', 0):+.2f}",
                    "--"
                )
                item = self.spread_positions_tree.insert("", "end", values=values)
                self.spread_positions_tree.item(item, tags=("active",))
                self.spread_positions_tree.tag_configure("active", foreground="blue")
            
            self.log_message("Spread positions refreshed")
            
        except Exception as e:
            self.log_message(f"Error refreshing spread positions: {e}")

    def update_spread_history_display(self):
        """Update the spread history treeview"""
        # Clear existing items
        for item in self.spread_history_tree.get_children():
            self.spread_history_tree.delete(item)
        
        # Add spread history items
        for spread in self.spread_orders[-20:]:  # Show last 20 spreads
            leg1_info = f"{spread['leg1']['transaction_type']} {spread['leg1']['tradingsymbol']}"
            leg2_info = f"{spread['leg2']['transaction_type']} {spread['leg2']['tradingsymbol']}"
            
            values = (
                spread["time"],
                spread["spread_type"],
                leg1_info,
                leg2_info,
                spread["leg1"]["quantity"],
                spread["status"]
            )
            
            item = self.spread_history_tree.insert("", "end", values=values)
            
            # Color code based on success
            if spread.get("success", False):
                self.spread_history_tree.item(item, tags=("success",))
            else:
                self.spread_history_tree.item(item, tags=("partial",))
        
        # Configure tags
        self.spread_history_tree.tag_configure("success", foreground="green")
        self.spread_history_tree.tag_configure("partial", foreground="orange")

    def update_spread_performance_metrics(self):
        """Update spread performance metrics"""
        try:
            total_spreads = len(self.spread_orders)
            successful_spreads = sum(1 for s in self.spread_orders if s.get("success", False))
            failed_spreads = total_spreads - successful_spreads
            
            # Calculate net P&L (simplified - in reality would track actual P&L)
            net_pnl = 0.0
            for spread in self.spread_orders:
                if spread.get("success", False):
                    # Add some simulated P&L based on spread type
                    if "ENTRY" in spread.get("spread_type", ""):
                        net_pnl += 50.0  # Simulated profit
                    elif "EXIT" in spread.get("spread_type", ""):
                        net_pnl += 30.0  # Simulated profit
            
            # Update labels
            self.total_spreads_label.config(text=str(total_spreads))
            self.successful_spreads_label.config(text=str(successful_spreads))
            self.failed_spreads_label.config(text=str(failed_spreads))
            
            # Color code net P&L
            if net_pnl > 0:
                pnl_color = 'green'
                pnl_text = f"₹{net_pnl:+.2f}"
            elif net_pnl < 0:
                pnl_color = 'red'
                pnl_text = f"₹{net_pnl:+.2f}"
            else:
                pnl_color = 'orange'
                pnl_text = "₹0.00"
            
            self.net_spread_pnl_label.config(text=pnl_text, foreground=pnl_color)
            
        except Exception as e:
            self.log_message(f"Error updating spread metrics: {e}")

    def clear_spread_history(self):
        """Clear the spread history"""
        if messagebox.askyesno("Confirm", "Are you sure you want to clear spread history?"):
            self.spread_orders = []
            self.update_spread_history_display()
            self.update_spread_performance_metrics()
            self.log_message("Spread history cleared")

    def test_spread_order(self):
        """Test spread order functionality without actually placing orders"""
        if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
            messagebox.showerror("Error", "Please load contracts first")
            return
        
        # Create a test spread record
        test_spread = {
            "time": datetime.now().strftime("%H:%M:%S"),
            "spread_id": f"TEST_{int(time.time())}",
            "spread_type": "TEST ENTRY Spread",
            "leg1": {
                "tradingsymbol": self.next_month_contract,
                "transaction_type": "BUY",
                "quantity": 1
            },
            "leg2": {
                "tradingsymbol": self.current_month_contract,
                "transaction_type": "SELL",
                "quantity": 1
            },
            "order_type": "MARKET",
            "product": "NRML",
            "price": "MARKET",
            "status": "TEST",
            "success": True
        }
        
        # Add to spread orders
        self.spread_orders.append(test_spread)
        
        # Update UI
        self.update_spread_history_display()
        self.update_spread_performance_metrics()
        
        messagebox.showinfo("Test Successful", "Test spread order recorded successfully!")

    # ==============================================
    # AUTO-EXIT FUNCTIONS
    # ==============================================

    def toggle_auto_exit(self):
        """Toggle auto-exit functionality"""
        self.auto_exit_enabled = self.auto_exit_var.get()
        status = "enabled" if self.auto_exit_enabled else "disabled"
        self.log_auto_exit(f"Auto exit {status}")

    def start_auto_exit_monitoring(self):
        """Start monitoring for auto-exit conditions"""
        if not self.is_logged_in:
            messagebox.showerror("Error", "Please login first")
            return
        
        try:
            # Get profit target and stop loss
            self.auto_exit_profit_target = float(self.profit_target_var.get())
            #self.auto_exit_stop_loss = float(self.stop_loss_var.get())
            
            # Update status labels
            self.profit_target_status.config(text=f"₹{self.auto_exit_profit_target:.2f}")
            #self.stop_loss_status.config(text=f"₹{self.auto_exit_stop_loss:.2f}")
            
            # Check if we have a position
            self.check_current_position()
            
            if not self.current_position:
                self.log_auto_exit("No active position found. Waiting for new position...")
            
            self.auto_exit_running = True
            
            # Start monitoring thread
            threading.Thread(target=self.monitor_for_auto_exit, daemon=True).start()
            
            #self.log_auto_exit(f"Auto exit monitoring started. Profit target: ₹{self.auto_exit_profit_target}, Stop loss: ₹{self.auto_exit_stop_loss}")
            self.log_auto_exit(f"Auto exit monitoring started. Profit target: ₹{self.auto_exit_profit_target}")
            
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numbers for profit target and stop loss")

    def stop_auto_exit_monitoring(self):
        """Stop auto-exit monitoring"""
        self.auto_exit_running = False
        self.log_auto_exit("Auto exit monitoring stopped")

    def check_current_position(self):
        """Check current positions and set up auto-exit"""
        try:
            # Get positions from Zerodha
            positions_data = self.kite.positions()
            day_positions = positions_data.get('day', [])
            net_positions = positions_data.get('net', [])
            
            all_positions = day_positions + net_positions
            
            # Find active MCX futures positions
            active_positions = []
            for position in all_positions:
                if (position['exchange'] == 'MCX' and 
                    position['instrument_type'] == 'FUT' and 
                    position['quantity'] != 0):
                    active_positions.append(position)
            
            if len(active_positions) > 0:
                # For simplicity, track the first position
                position = active_positions[0]
                self.current_position = {
                    'tradingsymbol': position['tradingsymbol'],
                    'quantity': position['quantity'],
                    'avg_price': position['average_price'],
                    'direction': 'BUY' if position['quantity'] > 0 else 'SELL'
                }
                
                # Set trade start price
                self.trade_start_price = position['average_price']
                self.trade_direction = self.current_position['direction']
                
                # Update UI
                self.update_position_display()
                
                self.log_auto_exit(f"Active position found: {self.current_position['tradingsymbol']} "
                                  f"{self.current_position['direction']} {abs(self.current_position['quantity'])} "
                                  f"@ ₹{self.trade_start_price:.2f}")
                
                return True
            else:
                self.current_position = None
                self.update_position_display()
                return False
                
        except Exception as e:
            self.log_auto_exit(f"Error checking positions: {e}")
            return False

    def update_position_display(self):
        """Update the position display in auto-exit tab"""
        if self.current_position:
            position_text = f"{self.current_position['tradingsymbol']} "
            position_text += f"{self.current_position['direction']} {abs(self.current_position['quantity'])}"
            
            self.position_status_label.config(text=position_text, foreground='green')
            self.entry_price_label.config(text=f"₹{self.trade_start_price:.2f}")
            self.exit_button.config(state='normal')
        else:
            self.position_status_label.config(text="No active position", foreground='red')
            self.entry_price_label.config(text="₹0.00")
            self.current_pnl_label.config(text="₹0.00")
            self.exit_button.config(state='disabled')

    def monitor_for_auto_exit(self):
        """Monitor positions for auto-exit conditions"""
        update_interval = 3  # seconds
        
        while self.auto_exit_running and self.is_logged_in:
            try:
                # Check current position
                if not self.current_position:
                    # Try to find a new position
                    self.check_current_position()
                    time.sleep(5)
                    continue
                
                # Get current price
                tradingsymbol = self.current_position['tradingsymbol']
                ltp_data = self.kite.ltp(f"MCX:{tradingsymbol}")
                current_price = ltp_data[f"MCX:{tradingsymbol}"]['last_price']
                
                # Calculate P&L
                if self.trade_direction == "BUY":
                    pnl = (current_price - self.trade_start_price) * abs(self.current_position['quantity'])
                else:  # SELL
                    pnl = (self.trade_start_price - current_price) * abs(self.current_position['quantity'])
                
                # Update P&L display
                self.root.after(0, lambda p=pnl: self.update_pnl_display(p))
                
                # Check exit conditions if auto-exit is enabled
                if self.auto_exit_enabled:
                    if pnl >= self.auto_exit_profit_target:
                        self.log_auto_exit(f"🎯 PROFIT TARGET REACHED: ₹{pnl:.2f}")
                        self.root.after(0, lambda: self.auto_exit_position("PROFIT"))
                    #disable stop loss for testing purpose    
                    # elif pnl <= -self.auto_exit_stop_loss:
                    #     self.log_auto_exit(f"⚠️ STOP LOSS HIT: ₹{pnl:.2f}")
                    #     self.root.after(0, lambda: self.auto_exit_position("STOP_LOSS"))
                
                time.sleep(update_interval)
                
            except Exception as e:
                self.log_auto_exit(f"Error in auto-exit monitoring: {e}")
                time.sleep(5)

    def update_pnl_display(self, pnl):
        """Update P&L display with color coding"""
        pnl_text = f"₹{pnl:+.2f}"
        
        if pnl > 0:
            self.current_pnl_label.config(text=pnl_text, foreground='green')
        elif pnl < 0:
            self.current_pnl_label.config(text=pnl_text, foreground='red')
        else:
            self.current_pnl_label.config(text=pnl_text, foreground='orange')

    def auto_exit_position(self, reason):
        """Automatically exit the current position"""
        if not self.current_position:
            self.log_auto_exit("No position to exit")
            return
        
        try:
            tradingsymbol = self.current_position['tradingsymbol']
            quantity = abs(self.current_position['quantity'])
            
            # Determine exit direction (opposite of entry)
            if self.trade_direction == "BUY":
                exit_transaction = "SELL"
            else:  # SELL
                exit_transaction = "BUY"
            
            # Execute exit order
            result = self.execute_order(
                tradingsymbol=tradingsymbol,
                transaction_type=exit_transaction,
                quantity=quantity,
                order_type="MARKET",
                product="NRML"
            )
            
            if result:
                # Get current price for P&L calculation
                ltp_data = self.kite.ltp(f"MCX:{tradingsymbol}")
                exit_price = ltp_data[f"MCX:{tradingsymbol}"]['last_price']
                
                # Calculate final P&L
                if self.trade_direction == "BUY":
                    final_pnl = (exit_price - self.trade_start_price) * quantity
                else:
                    final_pnl = (self.trade_start_price - exit_price) * quantity
                
                # Log the exit
                reason_text = {
                    "PROFIT": "Profit target achieved",
                    "STOP_LOSS": "Stop loss triggered"
                }.get(reason, "Auto exit")
                
                self.log_auto_exit(f"✅ AUTO EXIT: {reason_text}")
                self.log_auto_exit(f"   Exit price: ₹{exit_price:.2f}")
                self.log_auto_exit(f"   Final P&L: ₹{final_pnl:+.2f}")
                
                # Show notification
                self.show_auto_exit_notification(reason, final_pnl)
                
                # Reset position
                self.current_position = None
                self.update_position_display()
                
                # Play sound notification
                self.play_exit_sound(reason)
                
            else:
                self.log_auto_exit(f"❌ Failed to auto-exit position")
                
        except Exception as e:
            self.log_auto_exit(f"❌ Error in auto-exit: {e}")

    def manual_exit_position(self):
        """Manually exit the current position"""
        if not self.current_position:
            messagebox.showinfo("No Position", "No active position to exit")
            return
        
        if messagebox.askyesno("Confirm Exit", "Are you sure you want to exit this position?"):
            self.auto_exit_position("MANUAL")

    def show_auto_exit_notification(self, reason, pnl):
        """Show notification when auto-exit occurs"""
        # Create notification window
        window = tk.Toplevel(self.root)
        window.title("💰 AUTO EXIT TRIGGERED")
        window.geometry("400x300")
        
        # Make window stay on top
        window.attributes('-topmost', True)
        
        # Set background color based on reason
        if reason == "PROFIT":
            bg_color = '#E8F5E9'  # Light green
            title = "🎯 PROFIT TARGET REACHED!"
            emoji = "💰"
        elif reason == "STOP_LOSS":
            bg_color = '#FFEBEE'  # Light red
            title = "⚠️ STOP LOSS TRIGGERED!"
            emoji = "🛑"
        else:
            bg_color = '#FFF3E0'  # Light orange
            title = "📤 POSITION EXITED"
            emoji = "📤"
        
        window.configure(bg=bg_color)
        
        # Center window
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')
        
        # Title
        ttk.Label(window, text=emoji, font=('Arial', 48), background=bg_color).pack(pady=10)
        ttk.Label(window, text=title, font=('Arial', 14, 'bold'), background=bg_color).pack(pady=5)
        
        # P&L display
        pnl_color = 'green' if pnl > 0 else 'red' if pnl < 0 else 'orange'
        pnl_text = f"Final P&L: ₹{pnl:+.2f}"
        ttk.Label(window, text=pnl_text, font=('Arial', 16, 'bold'), 
                 foreground=pnl_color, background=bg_color).pack(pady=10)
        
        # Position info
        if self.current_position:
            info_text = f"{self.current_position['tradingsymbol']}\n"
            info_text += f"Entry: ₹{self.trade_start_price:.2f}\n"
            info_text += f"Quantity: {abs(self.current_position['quantity'])}"
            ttk.Label(window, text=info_text, font=('Arial', 10), 
                     background=bg_color).pack(pady=10)
        
        # Close button
        ttk.Button(window, text="OK", command=window.destroy).pack(pady=20)
        
        # Auto-close after 10 seconds
        window.after(10000, window.destroy)

    def play_exit_sound(self, reason):
        """Play sound notification for exit"""
        try:
            # Play system beep with different patterns
            if reason == "PROFIT":
                # Success beep pattern
                for i in range(3):
                    self.root.bell()
                    time.sleep(0.1)
            elif reason == "STOP_LOSS":
                # Warning beep pattern
                for i in range(2):
                    self.root.bell()
                    time.sleep(0.3)
                    self.root.bell()
                    time.sleep(0.1)
            else:
                # Regular beep
                self.root.bell()
        except:
            pass

    def log_auto_exit(self, message):
        """Add message to auto-exit log"""
        def update_log():
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.auto_exit_log.insert(tk.END, f"[{timestamp}] {message}\n")
            self.auto_exit_log.see(tk.END)
            # Also log to main log
            self.log_message(f"AUTO EXIT: {message}")
        
        self.root.after(0, update_log)

    # ==============================================
    # AUTO-TRADING SPREADS BASED ON SIGNALS
    # ==============================================

    def auto_trade_spread_based_on_signal(self, price_difference, signal_type):
        """Automatically trade calendar spreads based on entry/exit signals"""
        if not self.auto_spread_var.get():
            return  # Auto-trading disabled
        
        try:
            # Get spread quantity
            quantity = int(self.spread_quantity_var.get())
            
            # Place spread based on signal
            if signal_type == "ENTRY":
                if price_difference < float(self.entry_threshold_var.get()):
                    self.log_message(f"🤖 AUTO: Placing ENTRY Spread (Price Diff: {price_difference:.2f})")
                    self.place_calendar_spread("ENTRY")
            elif signal_type == "EXIT":
                if price_difference > float(self.exit_threshold_var.get()):
                    self.log_message(f"🤖 AUTO: Placing EXIT Spread (Price Diff: {price_difference:.2f})")
                    self.place_calendar_spread("EXIT")
                    
        except Exception as e:
            self.log_message(f"❌ Auto-trade failed: {e}")

    # ==============================================
    # MODIFIED ENTRY/EXIT POPUP WITH SPREAD BUTTONS
    # ==============================================

    def show_entry_exit_popup(self, price_difference, signal_type):
        """Show entry/exit popup based on price difference with spread trading options"""
        # Close existing popup if open
        if self.entry_exit_popup and self.entry_exit_popup.winfo_exists():
            self.entry_exit_popup.destroy()
        
        # Create new popup window with resizable panes
        window = tk.Toplevel(self.root)
        
        # Set window properties based on signal type
        if signal_type == "ENTRY":
            window.title("🎯 ENTRY SIGNAL - Consider Buying")
            smiley = "😊"
            message = "ENTRY SIGNAL - Consider BUYING"
            bg_color = '#E8F5E9'  # Light green
            text_color = 'dark green'
            urgency = "🔥 STRONG BUY SIGNAL"
        else:  # EXIT
            window.title("🚪 EXIT SIGNAL - Consider Selling")
            smiley = "😢"
            message = "EXIT SIGNAL - Consider SELLING"
            bg_color = '#FFEBEE'  # Light red
            text_color = 'dark red'
            urgency = "⚠️ STRONG SELL SIGNAL"
        
        window.geometry("800x600")
        
        # Make window resizable and draggable
        window.resizable(True, True)
        window.attributes('-topmost', True)
        window.focus_force()
        
        # Play system beep (multiple times for urgency)
        for _ in range(3):
            window.bell()
            time.sleep(0.1)
        
        # Store reference
        self.entry_exit_popup = window
        
        # Set urgent color
        window.configure(bg=bg_color)
        
        # Center window
        self.center_window(window)
        
        # Create main paned window for resizable sections
        main_pane = PanedWindow(window, orient=tk.VERTICAL, sashrelief=tk.RAISED, sashwidth=5)
        main_pane.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Top section: Signal header
        header_frame = ttk.Frame(main_pane)
        main_pane.add(header_frame)
        
        smiley_label = tk.Label(header_frame, text=smiley, font=('Arial', 72), bg=bg_color)
        smiley_label.pack(pady=5)
        
        urgency_label = ttk.Label(header_frame, 
                                 text=urgency,
                                 font=('Arial', 18, 'bold'),
                                 foreground=text_color)
        urgency_label.pack(pady=5)
        
        signal_label = ttk.Label(header_frame,
                                text=message,
                                font=('Arial', 16, 'bold'),
                                foreground=text_color)
        signal_label.pack(pady=5)
        
        # Middle section: Signal details
        details_frame = ttk.LabelFrame(main_pane, text="Signal Details")
        main_pane.add(details_frame)
        
        # Create a grid for details
        details_grid = ttk.Frame(details_frame)
        details_grid.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Price Difference
        ttk.Label(details_grid, text="Price Difference (₹):", font=('Arial', 12, 'bold')).grid(row=0, column=0, sticky='w', pady=10)
        price_diff_label = ttk.Label(details_grid,
                                    text=f"{price_difference:+.2f}",
                                    font=('Arial', 14, 'bold'),
                                    foreground='green' if price_difference > 0 else 'red')
        price_diff_label.grid(row=0, column=1, sticky='w', pady=10, padx=10)
        
        # Threshold info
        ttk.Label(details_grid, text="Trigger Threshold:", font=('Arial', 11)).grid(row=1, column=0, sticky='w', pady=5)
        if signal_type == "ENTRY":
            threshold_text = f"Less than {self.entry_threshold}"
            threshold_color = 'red'
        else:
            threshold_text = f"More than {self.exit_threshold}"
            threshold_color = 'green'
        
        threshold_label = ttk.Label(details_grid,
                                   text=threshold_text,
                                   font=('Arial', 11, 'bold'),
                                   foreground=threshold_color)
        threshold_label.grid(row=1, column=1, sticky='w', pady=5, padx=10)
        
        # Contract names
        ttk.Label(details_grid, text="Current Contract:", font=('Arial', 10)).grid(row=2, column=0, sticky='w', pady=5)
        ttk.Label(details_grid, text=self.current_month_contract, font=('Arial', 10)).grid(row=2, column=1, sticky='w', pady=5, padx=10)
        
        ttk.Label(details_grid, text="Next Contract:", font=('Arial', 10)).grid(row=3, column=0, sticky='w', pady=5)
        ttk.Label(details_grid, text=self.next_month_contract, font=('Arial', 10)).grid(row=3, column=1, sticky='w', pady=5, padx=10)
        
        # Time of trigger
        trigger_time = datetime.now().strftime("%H:%M:%S")
        ttk.Label(details_grid, text="Signal Time:", font=('Arial', 9)).grid(row=4, column=0, sticky='w', pady=5)
        ttk.Label(details_grid, text=trigger_time, font=('Arial', 9)).grid(row=4, column=1, sticky='w', pady=5, padx=10)
        
        # Bottom section: Action buttons
        action_frame = ttk.LabelFrame(main_pane, text="Trading Actions")
        main_pane.add(action_frame, stretch="always")
        
        # Create horizontal paned window for buttons
        button_pane = PanedWindow(action_frame, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, sashwidth=5)
        button_pane.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Left side - Single leg trading
        single_leg_frame = ttk.LabelFrame(button_pane, text="Single Leg Trading")
        button_pane.add(single_leg_frame, stretch="always")
        
        if signal_type == "ENTRY":
            ttk.Button(single_leg_frame, text="BUY Current Month",
                      command=lambda: self.quick_buy(self.current_month_contract),
                      style="Buy.TButton").pack(pady=5, padx=10)
            
            ttk.Button(single_leg_frame, text="BUY Next Month",
                      command=lambda: self.quick_buy(self.next_month_contract),
                      style="Buy.TButton").pack(pady=5, padx=10)
            
            ttk.Button(single_leg_frame, text="BUY TOGETHER",
                      command=self.place_buy_together_order,
                      style="BuyTogether.TButton").pack(pady=5, padx=10)
        
        else:  # EXIT signal
            ttk.Button(single_leg_frame, text="SELL Current Month",
                      command=lambda: self.quick_sell(self.current_month_contract),
                      style="Sell.TButton").pack(pady=5, padx=10)
            
            ttk.Button(single_leg_frame, text="SELL Next Month",
                      command=lambda: self.quick_sell(self.next_month_contract),
                      style="Sell.TButton").pack(pady=5, padx=10)
        
        # Right side - Spread trading
        spread_frame = ttk.LabelFrame(button_pane, text="Spread Trading")
        button_pane.add(spread_frame, stretch="always")
        
        if signal_type == "ENTRY":
            ttk.Button(spread_frame, text="🎯 Place ENTRY Spread",
                      command=lambda: self.place_calendar_spread("ENTRY"),
                      style="EntrySpread.TButton").pack(pady=5, padx=10)
        else:
            ttk.Button(spread_frame, text="🚪 Place EXIT Spread",
                      command=lambda: self.place_calendar_spread("EXIT"),
                      style="ExitSpread.TButton").pack(pady=5, padx=10)
        
        # Auto-trade checkbox
        auto_trade_var = tk.BooleanVar(value=self.auto_spread_var.get())
        auto_check = ttk.Checkbutton(spread_frame, text="Auto-trade future signals",
                                    variable=auto_trade_var,
                                    command=lambda: self.toggle_auto_trading(auto_trade_var))
        auto_check.pack(pady=10)
        
        # Control buttons at bottom
        control_frame = ttk.Frame(action_frame)
        control_frame.pack(fill='both', expand=True,  pady=10)
        
        ttk.Button(control_frame, text="Show Detailed Analysis",
                  command=self.show_price_difference_popup).pack(side='left', padx=5)
        
        ttk.Button(control_frame, text="Show Comparison",
                  command=self.show_comparison_popup).pack(side='left', padx=5)
        
        ttk.Button(control_frame, text="Acknowledge Signal",
                  command=lambda: self.acknowledge_entry_exit_signal(window, signal_type)).pack(side='right', padx=5)
        
        ttk.Button(control_frame, text=f"Mute for {self.entry_exit_cooldown//60} min",
                  command=lambda: self.mute_entry_exit_signals(window)).pack(side='right', padx=5)
        
        # Log this signal
        self.log_message(f"🚨 {signal_type} SIGNAL: Price difference {price_difference:+.2f} (Threshold: {self.entry_threshold if signal_type == 'ENTRY' else self.exit_threshold})")
        
        # Check if we should auto-trade spread
        if self.auto_spread_var.get():
            self.root.after(1000, lambda: self.auto_trade_spread_based_on_signal(price_difference, signal_type))
        
        # Update last trigger time
        self.last_entry_exit_trigger_time = time.time()
        
        # Update signal display in main window
        self.update_signal_display(signal_type, price_difference)
        
        # Handle window close
        window.protocol("WM_DELETE_WINDOW", lambda: self.acknowledge_entry_exit_signal(window, signal_type))
        
        # Flash the window for attention
        self.flash_window(window, signal_type, 5)

    def toggle_auto_trading(self, var):
        """Toggle auto-trading of spreads"""
        self.auto_spread_var.set(var.get())
        status = "enabled" if var.get() else "disabled"
        self.log_message(f"🤖 Auto-spread trading {status}")

    def quick_buy(self, contract):
        """Quick buy function for entry/exit popup"""
        try:
            quantity = int(self.quantity_var.get())
            self.execute_order(
                tradingsymbol=contract,
                transaction_type="BUY",
                quantity=quantity,
                order_type="MARKET",
                product="NRML"
            )
        except Exception as e:
            messagebox.showerror("Error", f"Quick buy failed: {e}")

    def quick_sell(self, contract):
        """Quick sell function for entry/exit popup"""
        try:
            quantity = int(self.quantity_var.get())
            self.execute_order(
                tradingsymbol=contract,
                transaction_type="SELL",
                quantity=quantity,
                order_type="MARKET",
                product="NRML"
            )
        except Exception as e:
            messagebox.showerror("Error", f"Quick sell failed: {e}")

    def flash_window(self, window, signal_type, times=5):
        """Flash window for attention"""
        def flash(count):
            if count > 0 and window.winfo_exists():
                current_color = window.cget('bg')
                if signal_type == "ENTRY":
                    flash_color = '#C8E6C9' if current_color == '#E8F5E9' else '#E8F5E9'
                else:
                    flash_color = '#FFCDD2' if current_color == '#FFEBEE' else '#FFEBEE'
                
                window.configure(bg=flash_color)
                window.after(200, lambda: flash(count-1))
            elif window.winfo_exists():
                # Restore original color
                window.configure(bg='#E8F5E9' if signal_type == "ENTRY" else '#FFEBEE')
        
        flash(times)

    def acknowledge_entry_exit_signal(self, window, signal_type):
        """Acknowledge and close entry/exit popup"""
        window.destroy()
        self.entry_exit_popup = None
        
        # Reset status after cooldown
        self.root.after(10000, lambda: None)

    def mute_entry_exit_signals(self, window):
        """Mute entry/exit signals for specified time"""
        try:
            minutes = int(self.entry_exit_cooldown_var.get())
            self.entry_exit_cooldown = minutes * 60
            self.last_entry_exit_trigger_time = time.time()
            
            # Close window
            window.destroy()
            self.entry_exit_popup = None
            
            self.log_message(f"🔕 Entry/Exit signals muted for {minutes} minutes")
            
            # Reset after cooldown
            self.root.after(minutes * 60 * 1000, lambda: self.reset_entry_exit_mute())
            
        except ValueError:
            messagebox.showerror("Error", "Please enter valid cooldown minutes")

    def reset_entry_exit_mute(self):
        """Reset entry/exit mute status"""
        try:
            self.entry_exit_cooldown = int(self.entry_exit_cooldown_var.get()) * 60
            self.log_message("🔔 Entry/Exit signals unmuted")
        except ValueError:
            self.entry_exit_cooldown = 60

    def update_signal_display(self, signal_type, price_difference):
        """Update the signal display in the main window"""
        if signal_type == "ENTRY":
            self.signal_display.config(text="😊", fg='green')
            self.signal_text.config(text=f"ENTRY SIGNAL\nPrice Diff: {price_difference:+.2f}", foreground='green')
            self.signal_display.configure(bg='#E8F5E9')
        else:  # EXIT
            self.signal_display.config(text="😢", fg='red')
            self.signal_text.config(text=f"EXIT SIGNAL\nPrice Diff: {price_difference:+.2f}", foreground='red')
            self.signal_display.configure(bg='#FFEBEE')
        
        # Reset after 30 seconds
        self.root.after(30000, lambda: self.reset_signal_display())

    def reset_signal_display(self):
        """Reset the signal display to default"""
        self.signal_display.config(text="--", fg='black')
        self.signal_text.config(text="No Signal", foreground='black')
        self.signal_display.configure(bg='white')

    def show_price_difference_popup(self):
        """Show popup with price difference in rupees"""
        if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
            messagebox.showerror("Error", "Please load contracts first")
            return
        
        # Close existing popup if open
        if self.price_diff_popup and self.price_diff_popup.winfo_exists():
            self.price_diff_popup.destroy()
        
        # Get current data
        try:
            contracts = [self.current_month_contract, self.next_month_contract]
            instruments = [f"MCX:{contract}" for contract in contracts]
            quote_data = self.kite.quote(instruments)
            
            current_price = quote_data[f"MCX:{self.current_month_contract}"]['last_price']
            next_price = quote_data[f"MCX:{self.next_month_contract}"]['last_price']
            
            # Get PREVIOUS DAY CLOSE prices
            current_prev = self.previous_day_close_prices.get(self.current_month_contract, current_price)
            next_prev = self.previous_day_close_prices.get(self.next_month_contract, next_price)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get current prices: {e}")
            return
        
        # Calculate changes in rupees
        current_change_rupees = current_price - current_prev
        next_change_rupees = next_price - next_prev
        
        # Calculate price difference using the formula:
        # change difference in rs = current month (Current Price - Previous Close) - next month (Current Price - Previous Close)
        price_difference = current_change_rupees - next_change_rupees
        
        # Create new window with resizable panes
        window = tk.Toplevel(self.root)
        window.title(f"💰 Price Difference - {self.month_commodity.get()}")
        window.geometry("800x700")
        
        # Make window resizable and stay on top
        window.resizable(True, True)
        window.attributes('-topmost', True)
        
        # Store reference
        self.price_diff_popup = window
        
        # Center window
        self.center_window(window)
        
        # Create main vertical paned window
        main_pane = PanedWindow(window, orient=tk.VERTICAL, sashrelief=tk.RAISED, sashwidth=5)
        main_pane.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Top section: Title and formula
        title_frame = ttk.Frame(main_pane)
        main_pane.add(title_frame, stretch="always")
        
        title_label = ttk.Label(title_frame, 
                               text=f"💰 Price Difference in Rupees", 
                               font=('Arial', 16, 'bold'))
        title_label.pack(pady=5)
        
        subtitle_label = ttk.Label(title_frame, 
                                  text=f"{self.month_commodity.get()} - Current vs Next Month",
                                  font=('Arial', 12))
        #subtitle_label.pack(pady=2)
        subtitle_label.pack(fill='both', expand=True, padx=5, pady=5)
        
        formula_label = ttk.Label(title_frame,
                                 text="Formula: Price Difference = (Current Month Change ₹) - (Next Month Change ₹)",
                                 font=('Arial', 10, 'italic'))
        #formula_label.pack(pady=5)
        formula_label.pack(fill='both', expand=True, padx=5, pady=5)
        

        
        self.price_diff_timestamp = ttk.Label(title_frame, 
                                            text=f"Last update: {datetime.now().strftime('%H:%M:%S')}",
                                            font=('Arial', 9))
        self.price_diff_timestamp.pack(pady=5)
        
        # Middle section: Price change details with horizontal panes
        details_pane = PanedWindow(main_pane, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, sashwidth=5)
        main_pane.add(details_pane, stretch="always")
        
        # Left pane: Current Month
        current_frame = ttk.LabelFrame(details_pane, text="Current Month")
        details_pane.add(current_frame)
        
        ttk.Label(current_frame, text=f"Contract:", font=('Arial', 11)).pack(pady=5)
        ttk.Label(current_frame, text=self.current_month_contract, 
                 font=('Arial', 10, 'bold')).pack(pady=2)
        
        ttk.Label(current_frame, text=f"Current Price:", font=('Arial', 10)).pack(pady=5)
        ttk.Label(current_frame, text=f"₹{current_price:.2f}", 
                 font=('Arial', 12)).pack(pady=2)
        
        ttk.Label(current_frame, text=f"Previous Close:", font=('Arial', 10)).pack(pady=5)
        ttk.Label(current_frame, text=f"₹{current_prev:.2f}", 
                 font=('Arial', 10)).pack(pady=2)
        
        ttk.Label(current_frame, text=f"Change in Rupees:", font=('Arial', 11, 'bold')).pack(pady=10)
        self.price_diff_popup_current = ttk.Label(current_frame, 
                                                 text=f"₹{current_change_rupees:+.2f}",
                                                 font=('Arial', 14, 'bold'))
        self.price_diff_popup_current.pack(pady=5)
        
        # Right pane: Next Month
        next_frame = ttk.LabelFrame(details_pane, text="Next Month")
        details_pane.add(next_frame, stretch="always")
        
        ttk.Label(next_frame, text=f"Contract:", font=('Arial', 11)).pack(pady=5)
        ttk.Label(next_frame, text=self.next_month_contract, 
                 font=('Arial', 10, 'bold')).pack(pady=2)
        
        ttk.Label(next_frame, text=f"Current Price:", font=('Arial', 10)).pack(pady=5)
        ttk.Label(next_frame, text=f"₹{next_price:.2f}", 
                 font=('Arial', 12)).pack(pady=2)
        
        ttk.Label(next_frame, text=f"Previous Close:", font=('Arial', 10)).pack(pady=5)
        ttk.Label(next_frame, text=f"₹{next_prev:.2f}", 
                 font=('Arial', 10)).pack(pady=2)
        
        ttk.Label(next_frame, text=f"Change in Rupees:", font=('Arial', 11, 'bold')).pack(pady=10)
        self.price_diff_popup_next = ttk.Label(next_frame, 
                                              text=f"₹{next_change_rupees:+.2f}",
                                              font=('Arial', 14, 'bold'))
        self.price_diff_popup_next.pack(pady=5)
        
        # Bottom section: Result and interpretation
        result_frame = ttk.LabelFrame(main_pane, text="Price Difference Result")
        main_pane.add(result_frame, stretch="always")
        
        # Formula display
        formula_text = f"Price Difference = (₹{current_change_rupees:+.2f}) - (₹{next_change_rupees:+.2f})"
        formula_display = ttk.Label(result_frame, text=formula_text, font=('Arial', 10))
        formula_display.pack(pady=5)
        
        # Result with color coding
        self.price_diff_popup_result = ttk.Label(result_frame, 
                                                text=f"Price Difference = ₹{price_difference:+.2f}",
                                                font=('Arial', 16, 'bold'))
        self.price_diff_popup_result.pack(pady=10)
        
        # Check entry/exit conditions
        should_trigger, signal_type, _ = self.check_entry_exit_condition(price_difference)
        
        if should_trigger:
            if signal_type == "ENTRY":
                signal_text = "🎯 ENTRY SIGNAL: Consider BUYING"
                signal_color = 'green'
                signal_bg = '#E8F5E9'
                advice = "Next month is performing significantly better"
            else:  # EXIT
                signal_text = "🚪 EXIT SIGNAL: Consider SELLING"
                signal_color = 'red'
                signal_bg = '#FFEBEE'
                advice = "Current month is performing significantly better"
            
            signal_label = ttk.Label(result_frame,
                                    text=signal_text,
                                    font=('Arial', 12, 'bold'),
                                    foreground=signal_color)
            signal_label.pack(pady=5)
            
            advice_label = ttk.Label(result_frame,
                                    text=advice,
                                    font=('Arial', 10, 'italic'),
                                    foreground=signal_color)
            advice_label.pack(pady=2)
            
            # Update window background
            window.configure(bg=signal_bg)
        
        # Action buttons
        button_frame = ttk.Frame(result_frame)
        button_frame.pack(pady=10)
        
        if price_difference < 0:  # Entry signal territory
            ttk.Button(button_frame, text="BUY Current Month",
                      command=lambda: self.quick_buy(self.current_month_contract),
                      style="Buy.TButton").pack(side='left', padx=2)
            
            ttk.Button(button_frame, text="BUY Next Month",
                      command=lambda: self.quick_buy(self.next_month_contract),
                      style="Buy.TButton").pack(side='left', padx=2)
            
            ttk.Button(button_frame, text="BUY TOGETHER",
                      command=self.place_buy_together_order,
                      style="BuyTogether.TButton").pack(side='left', padx=2)
        
        else:  # Exit signal territory or neutral
            ttk.Button(button_frame, text="SELL Current Month",
                      command=lambda: self.quick_sell(self.current_month_contract),
                      style="Sell.TButton").pack(side='left', padx=2)
            
            ttk.Button(button_frame, text="SELL Next Month",
                      command=lambda: self.quick_sell(self.next_month_contract),
                      style="Sell.TButton").pack(side='left', padx=2)
        
        ttk.Button(button_frame, text="Close", 
                  command=lambda: self.on_price_diff_popup_close(window)).pack(side='right', padx=5)
        
        ttk.Button(button_frame, text="Show Full Comparison", 
                  command=self.show_comparison_popup).pack(side='right', padx=5)
        
        # Handle window close
        window.protocol("WM_DELETE_WINDOW", lambda: self.on_price_diff_popup_close(window))
        
        # Start updates
        self.start_price_diff_popup_updates(window)

    def start_price_diff_popup_updates(self, window):
        """Start updating price difference popup window"""
        def update_popup():
            if not window.winfo_exists():
                return
            
            try:
                # Get current prices
                contracts = [self.current_month_contract, self.next_month_contract]
                instruments = [f"MCX:{contract}" for contract in contracts]
                quote_data = self.kite.quote(instruments)
                
                current_price = quote_data[f"MCX:{self.current_month_contract}"]['last_price']
                next_price = quote_data[f"MCX:{self.next_month_contract}"]['last_price']
                
                # Get PREVIOUS DAY CLOSE prices
                current_prev = self.previous_day_close_prices.get(self.current_month_contract, current_price)
                next_prev = self.previous_day_close_prices.get(self.next_month_contract, next_price)
                
                # Calculate changes in rupees
                current_change_rupees = current_price - current_prev
                next_change_rupees = next_price - next_prev
                
                # Calculate price difference using the formula:
                # change difference in rs = current month (Current Price - Previous Close) - next month (Current Price - Previous Close)
                price_difference = current_change_rupees - next_change_rupees
                
                # Update timestamp
                self.price_diff_timestamp.config(text=f"Last update: {datetime.now().strftime('%H:%M:%S')}")
                
                # Update current month change
                current_color = 'green' if current_change_rupees >= 0 else 'red'
                self.price_diff_popup_current.config(
                    text=f"₹{current_change_rupees:+.2f}",
                    foreground=current_color
                )
                
                # Update next month change
                next_color = 'green' if next_change_rupees >= 0 else 'red'
                self.price_diff_popup_next.config(
                    text=f"₹{next_change_rupees:+.2f}",
                    foreground=next_color
                )
                
                # Update result
                result_color = 'green' if price_difference > 0 else 'red' if price_difference < 0 else 'orange'
                self.price_diff_popup_result.config(
                    text=f"Price Difference = ₹{price_difference:+.2f}",
                    foreground=result_color
                )
                
            except Exception as e:
                print(f"Error updating price difference popup: {e}")
            
            # Schedule next update
            if window.winfo_exists():
                window.after(2000, update_popup)
        
        # Start updates
        window.after(1000, update_popup)

    def on_price_diff_popup_close(self, window):
        """Handle price difference popup window close"""
        window.destroy()
        self.price_diff_popup = None

    def log_message(self, message):
        """Add message to log"""
        def update_log():
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
            self.log_text.see(tk.END)
        
        self.root.after(0, update_log)

    def init_daily_performance_db(self):
        """Initialize SQLite database for daily performance tracking"""
        try:
            conn = sqlite3.connect(self.daily_performance_db)
            cursor = conn.cursor()
            
            # Create table for daily performance
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS daily_performance (
                    date DATE,
                    commodity TEXT,
                    current_month_contract TEXT,
                    next_month_contract TEXT,
                    current_month_close REAL,
                    next_month_close REAL,
                    current_performance REAL,
                    next_performance REAL,
                    relative_performance REAL,
                    smiley_status TEXT,
                    total_sum REAL,
                    PRIMARY KEY (date, commodity)
                )
            ''')
            
            # Create table for previous day closes
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS previous_day_closes (
                    date DATE,
                    contract_symbol TEXT,
                    close_price REAL,
                    volume INTEGER,
                    PRIMARY KEY (date, contract_symbol)
                )
            ''')
            
            # Create table for order history
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS order_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp DATETIME,
                    contract TEXT,
                    transaction_type TEXT,
                    quantity INTEGER,
                    price REAL,
                    order_type TEXT,
                    product TEXT,
                    status TEXT,
                    order_id TEXT,
                    remarks TEXT
                )
            ''')
            
            conn.commit()
            conn.close()
            self.log_message("Daily performance database initialized")
        except Exception as e:
            self.log_message(f"Error initializing database: {e}")

    def generate_login_url(self):
        """Generate login URL for Zerodha"""
        try:
            self.api_key = self.api_key_entry.get()
            if not self.api_key:
                messagebox.showerror("Error", "Please enter API Key")
                return
                
            self.kite = KiteConnect(api_key=self.api_key)
            
            login_url = self.kite.login_url()
            webbrowser.open(login_url)
            messagebox.showinfo("Login URL", f"Login URL generated and opened in browser.\nIf not, copy this URL:\n{login_url}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate login URL: {e}")

    def manual_login(self):
        """Manual login with request token"""
        try:
            self.api_key = self.api_key_entry.get()
            api_secret = self.api_secret_entry.get()
            request_token = self.request_token_entry.get()
            
            if not all([self.api_key, api_secret, request_token]):
                messagebox.showerror("Error", "Please fill all fields")
                return
            
            self.kite = KiteConnect(api_key=self.api_key)
            data = self.kite.generate_session(request_token, api_secret=api_secret)
            self.access_token = data['access_token']
            self.kite.set_access_token(self.access_token)
            
            # Save credentials
            self.save_credentials()
            
            self.is_logged_in = True
            self.login_status.config(text="Logged In Successfully", foreground='green')
            
            # Load instruments
            self.load_instruments()
            
            messagebox.showinfo("Success", "Login successful!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Login failed: {e}")

    def auto_login(self):
        """Auto login with saved credentials"""
        try:
            if not hasattr(self, 'api_key') or not self.api_key or not hasattr(self, 'access_token') or not self.access_token:
                messagebox.showerror("Error", "No saved credentials found")
                return
            
            self.kite = KiteConnect(api_key=self.api_key)
            self.kite.set_access_token(self.access_token)
            
            # Test connection
            profile = self.kite.profile()
            
            self.is_logged_in = True
            self.login_status.config(text=f"Auto Login Successful - {profile['user_name']}", foreground='green')
            
            # Load instruments
            self.load_instruments()
            
            messagebox.showinfo("Success", f"Auto login successful! Welcome {profile['user_name']}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Auto login failed: {e}")

    def load_instruments(self):
        """Load MCX instruments"""
        try:
            if self.kite and self.is_logged_in:
                # Get all instruments
                all_instruments = self.kite.instruments("MCX")
                self.instruments_df = pd.DataFrame(all_instruments)
                
                # Convert expiry to datetime if it's string
                if 'expiry' in self.instruments_df.columns and self.instruments_df['expiry'].dtype == 'object':
                    self.instruments_df['expiry'] = pd.to_datetime(self.instruments_df['expiry']).dt.date
                
                self.log_message(f"Loaded {len(self.instruments_df)} MCX instruments")
                
        except Exception as e:
            self.log_message(f"Error loading instruments: {e}")

    def get_monthly_contracts(self, base_symbol):
        """Get current and next month contracts"""
        try:
            if self.instruments_df is None:
                self.load_instruments()
                if self.instruments_df is None:
                    return []
            
            # Filter instruments for the base symbol (futures)
            relevant_instruments = self.instruments_df[
                (self.instruments_df['tradingsymbol'].str.startswith(base_symbol)) &
                (self.instruments_df['instrument_type'] == 'FUT')
            ].copy()
            
            if relevant_instruments.empty:
                self.log_message(f"No FUT contracts found for {base_symbol}")
                return []
            
            # Sort by expiry
            relevant_instruments = relevant_instruments.sort_values('expiry')
            
            # Get current date
            current_date = datetime.now().date()
            
            # Filter out expired contracts
            relevant_instruments = relevant_instruments[relevant_instruments['expiry'] >= current_date]
            
            # Get nearest 2 contracts (current and next month)
            if len(relevant_instruments) >= 2:
                selected_contracts = relevant_instruments.head(2)['tradingsymbol'].tolist()
            else:
                selected_contracts = relevant_instruments['tradingsymbol'].tolist()
            
            self.log_message(f"Found {len(selected_contracts)} contracts for {base_symbol}")
            return selected_contracts
            
        except Exception as e:
            self.log_message(f"Error getting monthly contracts: {str(e)}")
            return []

    def fetch_previous_day_closes(self):
        """Fetch previous day closing prices for the contracts"""
        if not self.is_logged_in:
            messagebox.showerror("Error", "Please login first")
            return
        
        if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
            messagebox.showerror("Error", "Please load contracts first")
            return
        
        try:
            # Get today and previous trading day
            today = datetime.now().date()
            
            # Try to get data for last 5 days to find a trading day
            for days_back in range(1, 6):
                check_date = today - timedelta(days=days_back)
                
                # Try to fetch historical data for previous day
                self.fetch_contract_historical_data(self.current_month_contract, check_date)
                self.fetch_contract_historical_data(self.next_month_contract, check_date)
                
                # Check if we got data for both contracts
                if (self.current_month_contract in self.previous_day_close_prices and 
                    self.next_month_contract in self.previous_day_close_prices):
                    break
            
            # Update display
            self.update_prev_close_display()
            
            current_prev = self.previous_day_close_prices.get(self.current_month_contract, "Not found")
            next_prev = self.previous_day_close_prices.get(self.next_month_contract, "Not found")
            
            messagebox.showinfo("Previous Day Close Fetched", 
                              f"Previous day closing prices fetched:\n"
                              f"Current Month: ₹{current_prev if isinstance(current_prev, (int, float)) else current_prev}\n"
                              f"Next Month: ₹{next_prev if isinstance(next_prev, (int, float)) else next_prev}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to fetch previous day closes: {e}")

    def fetch_contract_historical_data(self, contract_symbol, date_to_check):
        """Fetch historical data for a specific contract and date"""
        try:
            # Get instrument token
            instrument_token = self.get_instrument_token(contract_symbol)
            if not instrument_token:
                self.log_message(f"Cannot find instrument token for {contract_symbol}")
                return None
            
            # Convert date to string format for Zerodha API
            from_date = date_to_check.strftime("%Y-%m-%d")
            to_date = date_to_check.strftime("%Y-%m-%d")
            
            # Fetch historical data
            historical_data = self.kite.historical_data(
                instrument_token=instrument_token,
                from_date=from_date,
                to_date=to_date,
                interval="day",
                continuous=False
            )
            
            if historical_data and len(historical_data) > 0:
                # Get the last day's closing price
                last_day_data = historical_data[-1]
                close_price = last_day_data['close']
                
                # Store in dictionary
                self.previous_day_close_prices[contract_symbol] = close_price
                
                # Also save to database
                self.save_previous_day_close_to_db(contract_symbol, date_to_check, close_price)
                
                return close_price
            
            return None
            
        except Exception as e:
            self.log_message(f"Error fetching historical data for {contract_symbol}: {e}")
            return None

    def get_instrument_token(self, tradingsymbol):
        """Get instrument token for a trading symbol"""
        try:
            if self.instruments_df is None:
                self.load_instruments()
            
            if self.instruments_df is not None:
                # Search for the contract
                contract = self.instruments_df[
                    (self.instruments_df['tradingsymbol'] == tradingsymbol)
                ]
                
                if not contract.empty:
                    return int(contract.iloc[0]['instrument_token'])
            
            # Fallback: try to fetch fresh data
            all_instruments = self.kite.instruments("MCX")
            for inst in all_instruments:
                if inst['tradingsymbol'] == tradingsymbol:
                    return int(inst['instrument_token'])
            
            self.log_message(f"Instrument token not found for {tradingsymbol}")
            return None
                
        except Exception as e:
            self.log_message(f"Error getting instrument token for {tradingsymbol}: {e}")
            return None

    def save_previous_day_close_to_db(self, contract_symbol, date_obj, close_price):
        """Save previous day close to database"""
        try:
            conn = sqlite3.connect(self.daily_performance_db)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT OR REPLACE INTO previous_day_closes 
                (date, contract_symbol, close_price)
                VALUES (?, ?, ?)
            ''', (date_obj, contract_symbol, close_price))
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            self.log_message(f"Error saving previous day close to DB: {e}")

    def update_prev_close_display(self):
        """Update previous day close price display"""
        if hasattr(self, 'current_prev_close_label') and hasattr(self, 'next_prev_close_label'):
            current_prev = self.previous_day_close_prices.get(self.current_month_contract, 0)
            next_prev = self.previous_day_close_prices.get(self.next_month_contract, 0)
            
            if current_prev:
                self.current_prev_close_label.config(text=f"Prev Close: ₹{current_prev:.2f}")
            else:
                self.current_prev_close_label.config(text="Prev Close: Not set")
            
            if next_prev:
                self.next_prev_close_label.config(text=f"Prev Close: ₹{next_prev:.2f}")
            else:
                self.next_prev_close_label.config(text="Prev Close: Not set")

    def set_manual_previous_close(self):
        """Set previous day close prices manually"""
        if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
            messagebox.showerror("Error", "Please load contracts first")
            return
        
        # Create dialog for manual entry
        dialog = tk.Toplevel(self.root)
        dialog.title("Set Manual Previous Day Close Prices")
        dialog.geometry("400x250")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Make dialog resizable
        dialog.resizable(True, True)
        
        ttk.Label(dialog, text=f"Enter Previous Day Close Prices:", 
                 font=('Arial', 10, 'bold')).pack(pady=10)
        
        # Current month price
        current_frame = ttk.Frame(dialog)
        current_frame.pack(fill='both', expand=True,  padx=20, pady=5)
        ttk.Label(current_frame, text=f"{self.current_month_contract}:").pack(side='left')
        current_price_entry = ttk.Entry(current_frame, width=15)
        current_price_entry.pack(side='left', padx=10)
        current_price_entry.insert(0, str(self.previous_day_close_prices.get(self.current_month_contract, '')))
        
        # Next month price
        next_frame = ttk.Frame(dialog)
        next_frame.pack(fill='both', expand=True,  padx=20, pady=5)
        ttk.Label(next_frame, text=f"{self.next_month_contract}:").pack(side='left')
        next_price_entry = ttk.Entry(next_frame, width=15)
        next_price_entry.pack(side='left', padx=10)
        next_price_entry.insert(0, str(self.previous_day_close_prices.get(self.next_month_contract, '')))
        
        def save_manual_prices():
            try:
                current_price = float(current_price_entry.get())
                next_price = float(next_price_entry.get())
                
                self.previous_day_close_prices[self.current_month_contract] = current_price
                self.previous_day_close_prices[self.next_month_contract] = next_price
                
                self.update_prev_close_display()
                dialog.destroy()
                
                messagebox.showinfo("Success", "Manual previous day closes set successfully")
                
            except ValueError:
                messagebox.showerror("Error", "Please enter valid numbers")
        
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="Save", command=save_manual_prices).pack(side='left', padx=10)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side='left', padx=10)

    def start_month_comparison(self):
        """Start month comparison monitoring using PREVIOUS DAY CLOSE"""
        if not self.is_logged_in:
            messagebox.showerror("Error", "Please login first")
            return
        
        if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
            messagebox.showerror("Error", "Please load contracts first")
            return
        
        # Check if we have previous day closes
        if (self.current_month_contract not in self.previous_day_close_prices or 
            self.next_month_contract not in self.previous_day_close_prices):
            
            response = messagebox.askyesno("Previous Day Close Missing", 
                                         "Previous day closing prices not set. Would you like to fetch them now?")
            if response:
                self.fetch_previous_day_closes()
            else:
                response2 = messagebox.askyesno("Set Manual", 
                                              "Would you like to set them manually?")
                if response2:
                    self.set_manual_previous_close()
                else:
                    return
        
        self.month_comparison_running = True
        self.start_month_btn.config(state='disabled')
        self.stop_month_btn.config(state='normal')
        self.month_status_label.config(text="Status: Monitoring", foreground='green')
        self.trigger_status_label.config(text="Trigger Status: Ready", foreground='green')
        
        # Start monitoring thread
        threading.Thread(target=self.monitor_month_comparison, daemon=True).start()
        
        self.log_message(f"Started month comparison monitoring (vs Previous Day Close)")

    def stop_month_comparison(self):
        """Stop month comparison monitoring"""
        self.month_comparison_running = False
        self.start_month_btn.config(state='normal')
        self.stop_month_btn.config(state='disabled')
        self.month_status_label.config(text="Status: Stopped", foreground='red')
        
        self.log_message("Stopped month comparison monitoring")

    def monitor_month_comparison(self):
        """Monitor and compare current vs next month contracts vs PREVIOUS DAY CLOSE"""
        update_interval = 2  # seconds
        
        while self.month_comparison_running and self.is_logged_in:
            try:
                contracts = [self.current_month_contract, self.next_month_contract]
                instruments = [f"MCX:{contract}" for contract in contracts]
                
                quote_data = self.kite.quote(instruments)
                
                current_prices = {}
                for contract in contracts:
                    price = quote_data[f"MCX:{contract}"]['last_price']
                    current_prices[contract] = price
                
                # Update GUI with current prices and comparisons vs PREVIOUS DAY CLOSE
                self.update_month_comparison_display(current_prices)
                
                # Update popup window if it exists
                if self.comparison_popup and self.comparison_popup.winfo_exists():
                    self.update_comparison_popup_display(
                        self.comparison_popup,
                        current_prices[self.current_month_contract],
                        current_prices[self.next_month_contract],
                        self.previous_day_close_prices.get(self.current_month_contract, 0),
                        self.previous_day_close_prices.get(self.next_month_contract, 0)
                    )
                
                # Update price difference popup if it exists
                if self.price_diff_popup and self.price_diff_popup.winfo_exists():
                    # Trigger update through the main thread
                    self.root.after(0, lambda: self.update_price_diff_display())
                
                time.sleep(update_interval)
                
            except Exception as e:
                self.log_message(f"Error in month comparison monitoring: {e}")
                time.sleep(5)

    def update_price_diff_display(self):
        """Update price difference display in the main window"""
        try:
            if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
                return
            
            # Get current prices
            contracts = [self.current_month_contract, self.next_month_contract]
            instruments = [f"MCX:{contract}" for contract in contracts]
            quote_data = self.kite.quote(instruments)
            
            current_price = quote_data[f"MCX:{self.current_month_contract}"]['last_price']
            next_price = quote_data[f"MCX:{self.next_month_contract}"]['last_price']
            
            # Get PREVIOUS DAY CLOSE prices
            current_prev = self.previous_day_close_prices.get(self.current_month_contract, current_price)
            next_prev = self.previous_day_close_prices.get(self.next_month_contract, next_price)
            
            # Calculate changes in rupees
            current_change_rupees = current_price - current_prev
            next_change_rupees = next_price - next_prev
            
            # Calculate price difference using the formula:
            # change difference in rs = current month (Current Price - Previous Close) - next month (Current Price - Previous Close)
            price_difference = current_change_rupees - next_change_rupees
            if print_debug:
                print("price_difference: ", price_difference)
            
            update_existing_file(price_difference)
            
            # Update labels with colors
            current_color = 'green' if current_change_rupees >= 0 else 'red'
            next_color = 'green' if next_change_rupees >= 0 else 'red'
            diff_color = 'green' if price_difference > 0 else 'red' if price_difference < 0 else 'orange'
            
            self.price_diff_current.config(
                text=f"₹{current_change_rupees:+.2f}",
                foreground=current_color
            )
            self.price_diff_next.config(
                text=f"₹{next_change_rupees:+.2f}",
                foreground=next_color
            )
            self.price_diff_total.config(
                text=f"₹{price_difference:+.2f}",
                foreground=diff_color
            )
            
        except Exception as e:
            print(f"Error updating price difference display: {e}")

    def update_month_comparison_display(self, current_prices):
        """Update month comparison display vs PREVIOUS DAY CLOSE"""
        if not self.root.winfo_exists():
            return
        
        def update_gui():
            try:
                current_price = current_prices.get(self.current_month_contract, 0)
                next_price = current_prices.get(self.next_month_contract, 0)
                
                # Get PREVIOUS DAY CLOSE prices
                current_prev_close = self.previous_day_close_prices.get(self.current_month_contract, current_price)
                next_prev_close = self.previous_day_close_prices.get(self.next_month_contract, next_price)
                
                # Calculate changes from PREVIOUS DAY CLOSE
                if current_prev_close > 0:
                    current_change = ((current_price - current_prev_close) / current_prev_close) * 100
                else:
                    current_change = 0
                
                if next_prev_close > 0:
                    next_change = ((next_price - next_prev_close) / next_prev_close) * 100
                else:
                    next_change = 0
                
                # Calculate total sum of changes
                total_sum = current_change + next_change
                
                # Calculate price changes in rupees
                current_change_rupees = current_price - current_prev_close
                next_change_rupees = next_price - next_prev_close
                price_difference = current_change_rupees - next_change_rupees
                
                # Update price labels
                self.current_price_label.config(text=f"Current: ₹{current_price:.2f}")
                self.next_price_label.config(text=f"Current: ₹{next_price:.2f}")
                
                # Update change labels with colors
                current_color = 'green' if current_change >= 0 else 'red'
                next_color = 'green' if next_change >= 0 else 'red'
                
                self.current_change_label.config(
                    text=f"Change: {current_change:+.2f}%",
                    foreground=current_color
                )
                self.next_change_label.config(
                    text=f"Change: {next_change:+.2f}%",
                    foreground=next_color
                )
                
                # Update price difference display
                current_rupee_color = 'green' if current_change_rupees >= 0 else 'red'
                next_rupee_color = 'green' if next_change_rupees >= 0 else 'red'
                diff_color = 'green' if price_difference > 0 else 'red' if price_difference < 0 else 'orange'
                
                self.price_diff_current.config(
                    text=f"₹{current_change_rupees:+.2f}",
                    foreground=current_rupee_color
                )
                self.price_diff_next.config(
                    text=f"₹{next_change_rupees:+.2f}",
                    foreground=next_rupee_color
                )
                self.price_diff_total.config(
                    text=f"₹{price_difference:+.2f}",
                    foreground=diff_color
                )
                
                # Update price difference display in the main window
                # This ensures the price difference updates during live monitoring
                self.update_price_diff_display()
                
                # Check entry/exit condition
                should_trigger, signal_type, _ = self.check_entry_exit_condition(price_difference)
                
                if should_trigger:
                    self.show_entry_exit_popup(price_difference, signal_type)
                
                # Update total changes summary section
                self.update_total_changes_summary(current_change, next_change, total_sum)
                
                # Check trigger condition for special popup
                should_perf_trigger, difference = self.check_trigger_condition(current_change, next_change)
                
                #Yogesh 
                # if should_perf_trigger:
                #     self.show_triggered_popup(current_change, next_change, difference)
                
                # Determine comparison logic
                next_increased = next_change > 0
                current_decreased = current_change < 0
                
                # Calculate relative performance
                relative_performance = next_change - current_change
                
                # Determine smiley
                smiley_status = "NEUTRAL"
                if next_increased and current_decreased:
                    # Best case: next month up, current month down
                    smiley = "😊"
                    smiley_color = 'green'
                    comparison_text = "📈 Next month UP, Current DOWN vs Prev Close"
                    result_color = 'green'
                    smiley_status = "POSITIVE"
                elif relative_performance > 0.2:  # Next month performing better by 0.2%
                    smiley = "😊"
                    smiley_color = 'green'
                    comparison_text = f"📈 Next month +{relative_performance:.2f}% better"
                    result_color = 'green'
                    smiley_status = "POSITIVE"
                elif relative_performance < -0.2:  # Current month performing better
                    smiley = "☹️"
                    smiley_color = 'red'
                    comparison_text = f"📉 Current month +{abs(relative_performance):.2f}% better"
                    result_color = 'red'
                    smiley_status = "NEGATIVE"
                else:
                    smiley = "😐"
                    smiley_color = 'orange'
                    comparison_text = "⚖️ Months similar performance vs Prev Close"
                    result_color = 'orange'
                    smiley_status = "NEUTRAL"
                
                # Update smiley and text
                self.month_smiley_label.config(text=smiley, fg=smiley_color)
                self.month_comparison_text.config(text=comparison_text, foreground=result_color)
                
                # Update result label
                self.month_result_label.config(
                    text=f"Comparison: Next month is {relative_performance:+.2f}% vs Current",
                    foreground=result_color
                )
                
                # Update trigger status
                if self.last_trigger_time:
                    time_since = int(time.time() - self.last_trigger_time)
                    cooldown_left = max(0, self.trigger_cooldown - time_since)
                    if cooldown_left > 0:
                        self.trigger_status_label.config(
                            text=f"Trigger Cooldown: {cooldown_left}s",
                            foreground='orange'
                        )
                    else:
                        self.trigger_status_label.config(
                            text="Trigger Status: Ready",
                            foreground='green'
                        )
                
                # Save daily performance to database (including total sum)
                commodity = self.month_commodity.get()
                self.save_daily_performance(
                    commodity, self.current_month_contract, self.next_month_contract,
                    current_price, next_price, current_change, next_change,
                    relative_performance, smiley_status, total_sum
                )
                
                # Update history display
                self.update_history_display(commodity)
                
            except Exception as e:
                print(f"Error updating month comparison display: {e}")
        
        self.root.after(0, update_gui)

    def update_total_changes_summary(self, current_change, next_change, total_sum):
        """Update the total changes summary section"""
        try:
            # Update individual changes
            current_color = 'green' if current_change >= 0 else 'red'
            next_color = 'green' if next_change >= 0 else 'red'
            
            self.total_current_change.config(
                text=f"{current_change:+.2f}%",
                foreground=current_color
            )
            self.total_next_change.config(
                text=f"{next_change:+.2f}%",
                foreground=next_color
            )
            
            # Update performance difference
            perf_diff = next_change - current_change
            perf_color = 'green' if perf_diff > 0 else 'red' if perf_diff < 0 else 'orange'
            self.total_perf_diff.config(
                text=f"{perf_diff:+.2f}%",
                foreground=perf_color
            )
            
            # Update total sum with color coding
            if total_sum > 10.0:
                total_color = 'dark green'
                total_emoji = "🚀"
            elif total_sum > 0.2:
                total_color = 'green'
                total_emoji = "📈"
            elif total_sum < -10.0:
                total_color = 'dark red'
                total_emoji = "⚠️"
            elif total_sum < -0.2:
                total_color = 'red'
                total_emoji = "📉"
            else:
                total_color = 'orange'
                total_emoji = "⚖️"
            
            self.total_sum_label.config(
                text=f"{total_emoji} {total_sum:+.2f}%",
                foreground=total_color,
                font=('Arial', 12, 'bold')
            )
            
        except Exception as e:
            print(f"Error updating total changes summary: {e}")

    def save_daily_performance(self, commodity, current_contract, next_contract, 
                              current_close, next_close, current_perf, next_perf, 
                              relative_perf, smiley_status, total_sum=None):
        """Save daily performance to database"""
        try:
            conn = sqlite3.connect(self.daily_performance_db)
            cursor = conn.cursor()
            
            today = date.today()
            
            cursor.execute('''
                INSERT OR REPLACE INTO daily_performance 
                (date, commodity, current_month_contract, next_month_contract,
                 current_month_close, next_month_close, current_performance,
                 next_performance, relative_performance, smiley_status, total_sum)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (today, commodity, current_contract, next_contract,
                  current_close, next_close, current_perf, next_perf,
                  relative_perf, smiley_status, total_sum))
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            self.log_message(f"Error saving daily performance: {e}")

    def get_historical_performance(self, commodity, days=7):
        """Get historical performance data"""
        try:
            conn = sqlite3.connect(self.daily_performance_db)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT date, current_performance, next_performance, 
                       relative_performance, smiley_status, total_sum
                FROM daily_performance 
                WHERE commodity = ?
                ORDER BY date DESC
                LIMIT ?
            ''', (commodity, days))
            
            results = cursor.fetchall()
            conn.close()
            
            return results
            
        except Exception as e:
            self.log_message(f"Error getting historical performance: {e}")
            return []

    def update_history_display(self, commodity):
        """Update historical performance display"""
        try:
            history_data = self.get_historical_performance(commodity, days=7)
            
            self.history_text.delete(1.0, tk.END)
            
            if not history_data:
                self.history_text.insert(tk.END, "No historical data available")
                return
            
            self.history_text.insert(tk.END, "Date       | Curr%  | Next%  | Rel%   | Total%  | Status\n")
            self.history_text.insert(tk.END, "-" * 60 + "\n")
            
            for record in history_data:
                date_str, curr_perf, next_perf, rel_perf, smiley, total_sum = record
                
                # Format date
                if isinstance(date_str, str):
                    display_date = date_str[:10]  # Take first 10 chars
                else:
                    display_date = str(date_str)[:10]
                
                # Format percentages
                curr_str = f"{curr_perf:+.1f}" if curr_perf is not None else "N/A"
                next_str = f"{next_perf:+.1f}" if next_perf is not None else "N/A"
                rel_str = f"{rel_perf:+.1f}" if rel_perf is not None else "N/A"
                total_str = f"{total_sum:+.1f}" if total_sum is not None else "N/A"
                
                # Add color tags based on smiley status
                line = f"{display_date} | {curr_str:6s} | {next_str:6s} | {rel_str:6s} | {total_str:7s} | {smiley}\n"
                
                self.history_text.insert(tk.END, line)
                
                # Apply colors based on total sum
                if total_sum is not None:
                    if total_sum > 10.0:
                        self.history_text.tag_add("dark_green", f"end-2l", f"end-1l")
                    elif total_sum > 0.2:
                        self.history_text.tag_add("green", f"end-2l", f"end-1l")
                    elif total_sum < -10.0:
                        self.history_text.tag_add("dark_red", f"end-2l", f"end-1l")
                    elif total_sum < -0.2:
                        self.history_text.tag_add("red", f"end-2l", f"end-1l")
                    else:
                        self.history_text.tag_add("orange", f"end-2l", f"end-1l")
                else:
                    self.history_text.tag_add("gray", f"end-2l", f"end-1l")
            
            # Configure text colors
            self.history_text.tag_config("dark_green", foreground="dark green")
            self.history_text.tag_config("green", foreground="green")
            self.history_text.tag_config("dark_red", foreground="dark red")
            self.history_text.tag_config("red", foreground="red")
            self.history_text.tag_config("orange", foreground="orange")
            self.history_text.tag_config("gray", foreground="gray")
            
        except Exception as e:
            self.log_message(f"Error updating history display: {e}")

    def center_window(self, window):
        """Center a window on screen"""
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'{width}x{height}+{x}+{y}')

    def test_connection(self):
        """Test connection to Zerodha"""
        if not self.is_logged_in:
            messagebox.showinfo("Not Logged In", "Please login first")
            return
        
        try:
            profile = self.kite.profile()
            self.log_message(f"Connection test successful! User: {profile['user_name']}")
            messagebox.showinfo("Success", f"Connected to Zerodha as {profile['user_name']}")
        except Exception as e:
            self.log_message(f"Connection test failed: {e}")
            messagebox.showerror("Error", f"Connection failed: {e}")

    # Missing methods from previous implementations that need to be included
    def check_trigger_condition(self, current_change, next_change):
        """
        Check if next month's performance is significantly better than current month's
        Returns: (bool, float difference)
        """
        try:
            # Update threshold from GUI
            self.trigger_threshold = float(self.trigger_threshold_var.get())
            self.trigger_cooldown = int(self.cooldown_var.get())
            
            # Calculate difference
            difference = next_change - current_change
            
            # Check if next month is performing significantly better
            if difference > self.trigger_threshold:
                # Check cooldown
                current_time = time.time()
                if self.last_trigger_time is None or (current_time - self.last_trigger_time) > self.trigger_cooldown:
                    return True, difference
            return False, difference
            
        except ValueError:
            # If invalid threshold, use defaults
            if next_change - current_change > 0.2:
                current_time = time.time()
                if self.last_trigger_time is None or (current_time - self.last_trigger_time) > 60:
                    return True, next_change - current_change
            return False, next_change - current_change

    def test_triggered_popup(self):
        """Test the triggered popup display"""
        if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
            messagebox.showerror("Error", "Please load contracts first")
            return
        
        # Simulate trigger condition
        current_change = -0.2  # Current month down 0.2%
        next_change = 1.5      # Next month up 1.5%
        difference = 2.0       # 2% difference
        
        self.show_triggered_popup(current_change, next_change, difference)

    def show_triggered_popup(self, current_change, next_change, difference):
        """Show triggered popup when next month is performing significantly better"""
        # Close existing popup if open
        if self.triggered_popup and self.triggered_popup.winfo_exists():
            self.triggered_popup.destroy()
        
        # Calculate total sum of changes
        total_sum = current_change + next_change
        
        # Create new popup window with resizable panes
        window = tk.Toplevel(self.root)
        window.title("🚨 ALERT: Next Month Outperforming!")
        window.geometry("700x600")
        
        # Make window resizable and draggable
        window.resizable(True, True)
        window.attributes('-topmost', True)
        window.focus_force()
        
        # Play system beep
        window.bell()
        
        # Store reference
        self.triggered_popup = window
        
        # Set urgent color
        window.configure(bg='#FFE5E5')  # Light red background
        
        # Center window
        self.center_window(window)
        
        # Create main vertical paned window
        main_pane = PanedWindow(window, orient=tk.VERTICAL, sashrelief=tk.RAISED, sashwidth=5)
        main_pane.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Top section: Header
        header_frame = ttk.Frame(main_pane)
        main_pane.add(header_frame, stretch="always")
        
        alert_label = tk.Label(header_frame, text="⚠️", font=('Arial', 48), bg='#FFE5E5')
        alert_label.pack(side='left', padx=10)
        
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side='left', fill='y', padx=10)
        
        title_label = ttk.Label(title_frame, 
                               text="NEXT MONTH OUTPERFORMING!",
                               font=('Arial', 16, 'bold'),
                               foreground='red')
        title_label.pack(pady=5)
        
        subtitle_label = ttk.Label(title_frame,
                                  text=f"{self.month_commodity.get()} - Month Performance Alert",
                                  font=('Arial', 12))
        subtitle_label.pack()
        
        # Middle section: Performance details with horizontal panes
        details_pane = PanedWindow(main_pane, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, sashwidth=5)
        main_pane.add(details_pane, stretch="always")
        
        # Left pane: Current Month
        current_frame = ttk.LabelFrame(details_pane, text="Current Month")
        details_pane.add(current_frame, stretch="always")
        
        ttk.Label(current_frame, text="Performance:", font=('Arial', 11)).pack(pady=5)
        current_perf_label = ttk.Label(current_frame, 
                                      text=f"{current_change:+.2f}%",
                                      font=('Arial', 14, 'bold'),
                                      foreground='green' if current_change >= 0 else 'red')
        current_perf_label.pack(pady=5)
        
        ttk.Label(current_frame, text="Contract:", font=('Arial', 10)).pack(pady=5)
        ttk.Label(current_frame, text=self.current_month_contract, font=('Arial', 10)).pack(pady=2)
        
        # Right pane: Next Month
        next_frame = ttk.LabelFrame(details_pane, text="Next Month")
        details_pane.add(next_frame, stretch="always")
        
        ttk.Label(next_frame, text="Performance:", font=('Arial', 11)).pack(pady=5)
        next_perf_label = ttk.Label(next_frame,
                                   text=f"{next_change:+.2f}%",
                                   font=('Arial', 14, 'bold'),
                                   foreground='green' if next_change >= 0 else 'red')
        next_perf_label.pack(pady=5)
        
        ttk.Label(next_frame, text="Contract:", font=('Arial', 10)).pack(pady=5)
        ttk.Label(next_frame, text=self.next_month_contract, font=('Arial', 10)).pack(pady=2)
        
        # Bottom section: Results and actions
        result_frame = ttk.LabelFrame(main_pane, text="Results and Actions")
        main_pane.add(result_frame, stretch="always")
        
        # Performance difference (highlighted)
        ttk.Label(result_frame, text="Performance Gap:", font=('Arial', 12, 'bold')).pack(pady=10)
        diff_label = ttk.Label(result_frame,
                              text=f"{difference:+.2f}%",
                              font=('Arial', 16, 'bold'),
                              foreground='green')
        diff_label.pack(pady=5)
        
        # TOTAL SUM of changes
        ttk.Label(result_frame, text="TOTAL SUM of Changes:", 
                 font=('Arial', 12, 'bold')).pack(pady=10)
        total_sum_label = ttk.Label(result_frame,
                                   text=f"{total_sum:+.2f}%",
                                   font=('Arial', 16, 'bold'),
                                   foreground='blue' if total_sum > 0 else 'red' if total_sum < 0 else 'orange')
        total_sum_label.pack(pady=5)
        
        # Action buttons
        button_frame = ttk.Frame(result_frame)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="BUY Next Month",
                  command=lambda: self.quick_buy(self.next_month_contract),
                  style="Buy.TButton").pack(side='left', padx=5)
        
        ttk.Button(button_frame, text="BUY TOGETHER",
                  command=self.place_buy_together_order,
                  style="BuyTogether.TButton").pack(side='left', padx=5)
        
        ttk.Button(button_frame, text="Show Detailed Comparison",
                  command=self.show_comparison_popup).pack(side='left', padx=5)
        
        ttk.Button(button_frame, text="Show Price Difference",
                  command=self.show_price_difference_popup).pack(side='left', padx=5)
        
        ttk.Button(button_frame, text="Acknowledge",
                  command=lambda: self.acknowledge_trigger(window)).pack(side='right', padx=5)
        
        ttk.Button(button_frame, text="Mute Alerts for 1 min",
                  command=lambda: self.mute_alerts(60, window)).pack(side='right', padx=5)
        
        # Log this trigger
        self.log_message(f"🚨 TRIGGER: Next month outperforming by {difference:.2f}% (Total: {total_sum:+.2f}%)")
        
        # Update trigger time
        self.last_trigger_time = time.time()
        
        # Handle window close
        window.protocol("WM_DELETE_WINDOW", lambda: self.acknowledge_trigger(window))

    def acknowledge_trigger(self, window):
        """Acknowledge and close triggered popup"""
        window.destroy()
        self.triggered_popup = None
        
        # Reset status after 10 seconds
        self.root.after(10000, lambda: None)

    def mute_alerts(self, seconds, window):
        """Mute alerts for specified number of seconds"""
        self.trigger_cooldown = seconds
        self.last_trigger_time = time.time()
        
        # Close window
        window.destroy()
        self.triggered_popup = None
        
        self.log_message(f"🔕 Alerts muted for {seconds} seconds")
        
        # Reset after cooldown
        self.root.after(seconds * 1000, lambda: self.reset_mute())

    def reset_mute(self):
        """Reset mute status"""
        try:
            self.trigger_cooldown = int(self.cooldown_var.get())
            self.log_message("🔔 Alerts unmuted")
        except ValueError:
            self.trigger_cooldown = 60

    def show_comparison_popup(self):
        """Show common popup with both Current and Next Month contract changes"""
        if not hasattr(self, 'current_month_contract') or not hasattr(self, 'next_month_contract'):
            messagebox.showerror("Error", "Please load contracts first")
            return
        
        # Close existing popup if open
        if self.comparison_popup and self.comparison_popup.winfo_exists():
            self.comparison_popup.destroy()
        
        # Create new window with resizable panes
        window = tk.Toplevel(self.root)
        window.title(f"📊 {self.month_commodity.get()} - Month Comparison")
        window.geometry("800x700")
        
        # Make window resizable
        window.resizable(True, True)
        window.attributes('-topmost', True)
        
        # Store reference
        self.comparison_popup = window
        
        # Center window
        self.center_window(window)
        
        # Create main vertical paned window
        main_pane = PanedWindow(window, orient=tk.VERTICAL, sashrelief=tk.RAISED, sashwidth=5)
        main_pane.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Top section: Title
        title_frame = ttk.Frame(main_pane)
        main_pane.add(title_frame, stretch="always")
        
        title_label = ttk.Label(title_frame, 
                               text=f"{self.month_commodity.get()} - Month Comparison", 
                               font=('Arial', 16, 'bold'))
        title_label.pack(pady=5)
        
        subtitle_label = ttk.Label(title_frame, 
                                  text="Changes from Previous Day Close",
                                  font=('Arial', 12))
        subtitle_label.pack(pady=2)
        
        self.popup_timestamp = ttk.Label(title_frame, 
                                        text=f"Last update: {datetime.now().strftime('%H:%M:%S')}",
                                        font=('Arial', 9))
        self.popup_timestamp.pack(pady=5)
        
        # Middle section: Contract comparison with horizontal panes
        contracts_pane = PanedWindow(main_pane, orient=tk.HORIZONTAL, sashrelief=tk.RAISED, sashwidth=5)
        main_pane.add(contracts_pane, stretch="always")
        
        # Left pane: Current Month
        current_frame = ttk.LabelFrame(contracts_pane, text="Current Month")
        contracts_pane.add(current_frame, stretch="always")
        
        ttk.Label(current_frame, text=self.current_month_contract, 
                 font=('Arial', 12, 'bold')).pack(pady=10)
        
        self.popup_current_price = ttk.Label(current_frame, text="₹--", 
                                            font=('Arial', 20, 'bold'))
        self.popup_current_price.pack(pady=5)
        
        # Previous Close
        prev_frame = ttk.Frame(current_frame)
        prev_frame.pack(fill='both', expand=True,  pady=5)
        ttk.Label(prev_frame, text="Prev Close:").pack(side='left')
        self.popup_current_prev = ttk.Label(prev_frame, text="₹--", 
                                           font=('Arial', 10))
        self.popup_current_prev.pack(side='left', padx=5)
        
        # Change in Rupees
        change_frame = ttk.LabelFrame(current_frame, text="Change (₹)")
        change_frame.pack(fill='both', expand=True,  pady=10, padx=10)
        
        self.popup_current_rupee_change = ttk.Label(change_frame, text="₹--", 
                                                   font=('Arial', 18, 'bold'))
        self.popup_current_rupee_change.pack(pady=10)
        
        # Percentage change
        self.popup_current_percent = ttk.Label(current_frame, text="(--%)", 
                                              font=('Arial', 14))
        self.popup_current_percent.pack(pady=5)
        
        # Status indicator
        self.popup_current_status = ttk.Label(current_frame, text="--", 
                                             font=('Arial', 12, 'bold'))
        self.popup_current_status.pack(pady=5)
        
        # Right pane: Next Month
        next_frame = ttk.LabelFrame(contracts_pane, text="Next Month")
        contracts_pane.add(next_frame, stretch="always")
        
        ttk.Label(next_frame, text=self.next_month_contract, 
                 font=('Arial', 12, 'bold')).pack(pady=10)
        
        self.popup_next_price = ttk.Label(next_frame, text="₹--", 
                                         font=('Arial', 20, 'bold'))
        self.popup_next_price.pack(pady=5)
        
        # Previous Close
        next_prev_frame = ttk.Frame(next_frame)
        next_prev_frame.pack(fill='both', expand=True,  pady=5)
        ttk.Label(next_prev_frame, text="Prev Close:").pack(side='left')
        self.popup_next_prev = ttk.Label(next_prev_frame, text="₹--", 
                                        font=('Arial', 10))
        self.popup_next_prev.pack(side='left', padx=5)
        
        # Change in Rupees
        next_change_frame = ttk.LabelFrame(next_frame, text="Change (₹)")
        next_change_frame.pack(fill='both', expand=True,  pady=10, padx=10)
        
        self.popup_next_rupee_change = ttk.Label(next_change_frame, text="₹--", 
                                                font=('Arial', 18, 'bold'))
        self.popup_next_rupee_change.pack(pady=10)
        
        # Percentage change
        self.popup_next_percent = ttk.Label(next_frame, text="(--%)", 
                                           font=('Arial', 14))
        self.popup_next_percent.pack(pady=5)
        
        # Status indicator
        self.popup_next_status = ttk.Label(next_frame, text="--", 
                                          font=('Arial', 12, 'bold'))
        self.popup_next_status.pack(pady=5)
        
        # Bottom section: Comparison results
        comparison_frame = ttk.LabelFrame(main_pane, text="Month Comparison Results")
        main_pane.add(comparison_frame, stretch="always")
        
        # Price Difference
        self.popup_price_diff = ttk.Label(comparison_frame, 
                                         text="Next month is ₹-- higher",
                                         font=('Arial', 12, 'bold'))
        self.popup_price_diff.pack(pady=5)
        
        # Performance Difference
        self.popup_perf_diff = ttk.Label(comparison_frame, 
                                        text="Performance difference: --%",
                                        font=('Arial', 11))
        self.popup_perf_diff.pack(pady=2)
        
        # Price Difference in Rupees
        self.popup_price_diff_rupees = ttk.Label(comparison_frame,
                                                text="Price Difference (₹): --",
                                                font=('Arial', 11))
        self.popup_price_diff_rupees.pack(pady=2)
        
        # TOTAL SUM of changes
        self.popup_total_sum = ttk.Label(comparison_frame,
                                        text="TOTAL SUM of Changes: --%",
                                        font=('Arial', 11, 'bold'))
        self.popup_total_sum.pack(pady=2)
        
        # Smiley indicator
        self.popup_smiley = tk.Label(comparison_frame, text="😐", 
                                    font=('Arial', 36))
        self.popup_smiley.pack(pady=5)
        
        # Status text
        self.popup_status_text = ttk.Label(comparison_frame, text="--", 
                                          font=('Arial', 11))
        self.popup_status_text.pack(pady=2)
        
        # Action buttons frame
        button_frame = ttk.Frame(comparison_frame)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="BUY Current", 
                  command=lambda: self.quick_buy(self.current_month_contract),
                  style="Buy.TButton").pack(side='left', padx=2)
        
        ttk.Button(button_frame, text="SELL Current", 
                  command=lambda: self.quick_sell(self.current_month_contract),
                  style="Sell.TButton").pack(side='left', padx=2)
        
        ttk.Button(button_frame, text="BUY Next", 
                  command=lambda: self.quick_buy(self.next_month_contract),
                  style="Buy.TButton").pack(side='left', padx=2)
        
        ttk.Button(button_frame, text="SELL Next", 
                  command=lambda: self.quick_sell(self.next_month_contract),
                  style="Sell.TButton").pack(side='left', padx=2)
        
        ttk.Button(button_frame, text="BUY TOGETHER", 
                  command=self.place_buy_together_order,
                  style="BuyTogether.TButton").pack(side='left', padx=2)
        
        ttk.Button(button_frame, text="Close", 
                  command=lambda: self.on_comparison_popup_close(window)).pack(side='right', padx=5)
        
        ttk.Button(button_frame, text="Show Price Difference Details",
                  command=self.show_price_difference_popup).pack(side='right', padx=5)
        
        # Handle window close
        window.protocol("WM_DELETE_WINDOW", lambda: self.on_comparison_popup_close(window))
        
        # Start updates
        self.start_comparison_popup_updates(window)

    def start_comparison_popup_updates(self, window):
        """Start updating comparison popup window"""
        def update_popup():
            if not window.winfo_exists():
                return
            
            try:
                # Get current prices
                contracts = [self.current_month_contract, self.next_month_contract]
                instruments = [f"MCX:{contract}" for contract in contracts]
                
                quote_data = self.kite.quote(instruments)
                
                current_price = quote_data[f"MCX:{self.current_month_contract}"]['last_price']
                next_price = quote_data[f"MCX:{self.next_month_contract}"]['last_price']
                
                # Get PREVIOUS DAY CLOSE prices
                current_prev = self.previous_day_close_prices.get(self.current_month_contract, current_price)
                next_prev = self.previous_day_close_prices.get(self.next_month_contract, next_price)
                
                # Update the popup display
                self.update_comparison_popup_display(window, current_price, next_price, current_prev, next_prev)
                
            except Exception as e:
                print(f"Error updating comparison popup: {e}")
            
            # Schedule next update
            if window.winfo_exists():
                window.after(2000, update_popup)
        
        # Start updates
        window.after(1000, update_popup)

    def update_comparison_popup_display(self, window, current_price, next_price, current_prev, next_prev):
        """Update comparison popup with all data"""
        try:
            # Calculate changes for Current Month
            current_change = current_price - current_prev
            current_percent = ((current_price - current_prev) / current_prev * 100) if current_prev > 0 else 0
            
            # Calculate changes for Next Month
            next_change = next_price - next_prev
            next_percent = ((next_price - next_prev) / next_prev * 100) if next_prev > 0 else 0
            
            # Calculate price difference between months
            price_diff = next_price - current_price
            perf_diff = next_percent - current_percent
            
            # Calculate price difference in rupees (Current Change - Next Change)
            price_diff_rupees = current_change - next_change
            
            # Calculate total sum of changes
            total_sum = current_percent + next_percent
            
            # Update timestamp
            self.popup_timestamp.config(text=f"Last update: {datetime.now().strftime('%H:%M:%S')}")
            
            # Update Current Month section
            self.update_contract_section(
                price=current_price,
                prev_price=current_prev,
                rupee_change=current_change,
                percent_change=current_percent,
                is_current=True
            )
            
            # Update Next Month section
            self.update_contract_section(
                price=next_price,
                prev_price=next_prev,
                rupee_change=next_change,
                percent_change=next_percent,
                is_current=False
            )
            
            # Update comparison section (including total sum and price difference)
            self.update_comparison_section(price_diff, perf_diff, current_percent, next_percent, 
                                         total_sum, current_change, next_change, price_diff_rupees)
            
        except Exception as e:
            print(f"Error updating comparison popup display: {e}")

    def update_contract_section(self, price, prev_price, rupee_change, percent_change, is_current=True):
        """Update a contract section in the popup"""
        try:
            # Determine colors and status
            if rupee_change > 0:
                price_color = 'green'
                status_text = "▲ UP"
                change_text = f"+₹{abs(rupee_change):.2f}"
            elif rupee_change < 0:
                price_color = 'red'
                status_text = "▼ DOWN"
                change_text = f"-₹{abs(rupee_change):.2f}"
            else:
                price_color = 'orange'
                status_text = "⏺ FLAT"
                change_text = "₹0.00"
            
            # Format percentage
            percent_text = f"({percent_change:+.2f}%)"
            
            if is_current:
                # Update Current Month widgets
                self.popup_current_price.config(
                    text=f"₹{price:,.2f}",
                    foreground=price_color
                )
                self.popup_current_prev.config(
                    text=f"₹{prev_price:,.2f}",
                    foreground='gray'
                )
                self.popup_current_rupee_change.config(
                    text=change_text,
                    foreground=price_color
                )
                self.popup_current_percent.config(
                    text=percent_text,
                    foreground=price_color
                )
                self.popup_current_status.config(
                    text=status_text,
                    foreground=price_color
                )
            else:
                # Update Next Month widgets
                self.popup_next_price.config(
                    text=f"₹{price:,.2f}",
                    foreground=price_color
                )
                self.popup_next_prev.config(
                    text=f"₹{prev_price:,.2f}",
                    foreground='gray'
                )
                self.popup_next_rupee_change.config(
                    text=change_text,
                    foreground=price_color
                )
                self.popup_next_percent.config(
                    text=percent_text,
                    foreground=price_color
                )
                self.popup_next_status.config(
                    text=status_text,
                    foreground=price_color
                )
                
        except Exception as e:
            print(f"Error updating contract section: {e}")

    def update_comparison_section(self, price_diff, perf_diff, current_percent, next_percent, 
                                total_sum, current_change_rupees, next_change_rupees, price_diff_rupees):
        """Update the comparison section in the popup"""
        try:
            # Determine colors for price difference
            if price_diff > 0:
                diff_color = 'green'
                diff_text = f"Next month is ₹{abs(price_diff):.2f} HIGHER"
            elif price_diff < 0:
                diff_color = 'red'
                diff_text = f"Next month is ₹{abs(price_diff):.2f} LOWER"
            else:
                diff_color = 'orange'
                diff_text = "Months are SAME PRICE"
            
            # Update price difference
            self.popup_price_diff.config(
                text=diff_text,
                foreground=diff_color
            )
            
            # Update performance difference
            perf_text = f"Performance difference: {perf_diff:+.2f}%"
            self.popup_perf_diff.config(
                text=perf_text,
                foreground=diff_color
            )
            
            # Update price difference in rupees
            price_diff_color = 'green' if price_diff_rupees > 0 else 'red' if price_diff_rupees < 0 else 'orange'
            price_diff_text = f"Price Difference (₹): {price_diff_rupees:+.2f}"
            self.popup_price_diff_rupees.config(
                text=price_diff_text,
                foreground=price_diff_color
            )
            
            # Update total sum of changes
            # Determine color for total sum
            if total_sum > 10.0:
                total_color = 'dark green'
                total_emoji = "🚀"
            elif total_sum > 0.2:
                total_color = 'green'
                total_emoji = "📈"
            elif total_sum < -10.0:
                total_color = 'dark red'
                total_emoji = "⚠️"
            elif total_sum < -0.2:
                total_color = 'red'
                total_emoji = "📉"
            else:
                total_color = 'orange'
                total_emoji = "⚖️"
            
            total_text = f"{total_emoji} TOTAL SUM of Changes: {total_sum:+.2f}%"
            self.popup_total_sum.config(
                text=total_text,
                foreground=total_color
            )
            
            # Determine smiley based on performance
            next_up = next_percent > 0
            current_down = current_percent < 0
            
            if next_up and current_down:
                # Best case: next month up, current month down
                smiley = "😊"
                smiley_color = 'green'
                status_text = "📈 Next UP, Current DOWN"
                bg_color = 'light green'
            elif perf_diff > 0.5:
                # Next month performing better by 0.5%
                smiley = "😊"
                smiley_color = 'green'
                status_text = f"📈 Next month +{perf_diff:.2f}% better"
                bg_color = 'light green'
            elif perf_diff < -0.5:
                # Current month performing better
                smiley = "☹️"
                smiley_color = 'red'
                status_text = f"📉 Current month +{abs(perf_diff):.2f}% better"
                bg_color = 'light coral'
            else:
                # Similar performance
                smiley = "😐"
                smiley_color = 'orange'
                status_text = "⚖️ Months similar performance"
                bg_color = 'light yellow'
            
            # Update smiley and status
            self.popup_smiley.config(
                text=smiley,
                fg=smiley_color
            )
            self.popup_status_text.config(
                text=status_text,
                foreground=smiley_color
            )
            
            # Update window background based on total sum
            if self.comparison_popup and self.comparison_popup.winfo_exists():
                if total_sum > 10.0:
                    self.comparison_popup.configure(bg='#E8F5E9')  # Very light green
                elif total_sum > 0.2:
                    self.comparison_popup.configure(bg='#F1F8E9')  # Light green
                elif total_sum < -10.0:
                    self.comparison_popup.configure(bg='#FFEBEE')  # Very light red
                elif total_sum < -0.2:
                    self.comparison_popup.configure(bg='#FFE5E5')  # Light red
                else:
                    self.comparison_popup.configure(bg='light yellow')
                
            # Visual effect for significant differences
            if abs(total_sum) > 7.0:
                current_bg = self.popup_smiley.cget('bg')
                self.popup_smiley.config(
                    bg='gold' if current_bg == 'SystemButtonFace' else 'SystemButtonFace'
                )
                if self.comparison_popup and self.comparison_popup.winfo_exists():
                    self.comparison_popup.after(500, 
                        lambda: self.popup_smiley.config(bg='SystemButtonFace'))
            
        except Exception as e:
            print(f"Error updating comparison section: {e}")

    def on_comparison_popup_close(self, window):
        """Handle comparison popup window close"""
        window.destroy()
        self.comparison_popup = None

def main():
    root = tk.Tk()
    app = ZerodhaTradingApp(root)
    create_initial_file()
    root.mainloop()
    

if __name__ == "__main__":
    main()